import openpyxl
from openpyxl.utils import get_column_letter

file_path = r"d:\Projects\logistika\Mahalliy_yonalish_Avtomat_Pustoy_Tizimi (1).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

# ====== SHEET 1 STYLES ======
ws1 = wb["ASOSIY REYTING & BALANS"]
print("=" * 80)
print("SHEET 1 STYLES")
print()

key_rows = [1, 2, 4, 5]
for r in key_rows:
    print(f"--- Row {r} ---")
    for cell in ws1[r]:
        if cell.value is not None:
            col = get_column_letter(cell.column)
            print(f"  {col}{r}: val={cell.value}")
            print(f"    font: name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}, color={cell.font.color}")
            print(f"    fill: patternType={cell.fill.patternType}, fgColor={cell.fill.fgColor}, bgColor={cell.fill.bgColor}")
            print(f"    align: horizontal={cell.alignment.horizontal}, vertical={cell.alignment.vertical}, wrapText={cell.alignment.wrapText}")
            if cell.border:
                print(f"    border: left={cell.border.left}, right={cell.border.right}, top={cell.border.top}, bottom={cell.border.bottom}")
            print()

# ====== SHEET 2 STYLES ======
ws2 = wb["30 KUNLIK MOLIYA VA REYSLAR"]
print("=" * 80)
print("SHEET 2 STYLES")
print()

key_rows = [1, 2, 4, 5, 6, 7, 8, 9, 10]
for r in key_rows:
    print(f"--- Row {r} ---")
    for cell in ws2[r]:
        if cell.value is not None:
            col = get_column_letter(cell.column)
            print(f"  {col}{r}: val={cell.value}")
            print(f"    font: name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}, color={cell.font.color}")
            print(f"    fill: patternType={cell.fill.patternType}, fgColor={cell.fill.fgColor}")
            print(f"    align: horizontal={cell.alignment.horizontal}, vertical={cell.alignment.vertical}, wrapText={cell.alignment.wrapText}")
            print()
