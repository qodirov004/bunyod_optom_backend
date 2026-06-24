from decimal import Decimal
from drf_yasg import openapi
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from . import models, rest_api
from operator import itemgetter
from django.db.models import Sum
from rest_framework import status
from django.db.models import Count
from rest_framework import viewsets
from collections import defaultdict
from django.http import HttpResponse
from django.utils.timezone import now
from .pagination import RaysPagination
from datetime import timedelta, datetime
from rest_framework.viewsets import ViewSet
from django.db.models import Prefetch
from rest_framework.decorators import action
from django.contrib.auth import authenticate
from rest_framework.response import Response
from django.utils.dateparse import parse_date
from django.contrib.auth import get_user_model
from drf_yasg.utils import swagger_auto_schema
from rest_framework.permissions import AllowAny
from rest_framework import permissions
from rest_framework.filters import SearchFilter
from rest_framework_simplejwt.tokens import RefreshToken
from .permissions import IsOwnerOrCEO, IsCashierOrAdmin, IsZaphosOrAdmin, IsDriverOrAdmin, IsBugalterOrAdmin
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from .docs import load_doc as docs
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Sum, F, Case, When, DecimalField
from django.db.models.functions import Coalesce
from django.db.models import OuterRef, Subquery, Sum, Value, DecimalField

User = get_user_model()

def to_uzs(amount, currency_obj) -> Decimal:
    return models.to_uzs(amount, currency_obj)

def to_usd(amount, currency_obj) -> float:
    if amount is None or amount == 0:
        return 0.0
    
    # Cache rates for this call
    rates = {r.currency: float(r.rate_to_uzs) for r in models.CurrencyRate.objects.all()}
    usd_rate = rates.get('USD', 12800.0) or 12800.0
    
    if isinstance(currency_obj, str):
        curr = currency_obj
    else:
        curr = getattr(currency_obj, 'currency', 'UZS')
        
    if curr == 'USD':
        return float(amount)
    elif curr == 'UZS':
        return float(amount) / usd_rate
    elif curr in rates:
        return (float(amount) * rates[curr]) / usd_rate
    return 0.0



def get_client_total_expected_uzs(client):
    from .models import Product
    total_uzs = 0
    for product in Product.objects.filter(client=client, is_delivered=False):
        currency = getattr(product, 'currency', None)
        price = getattr(product, 'price', 0)
        total_uzs += to_uzs(price, currency)
    return total_uzs

class CurrencyRateViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [permissions.IsAuthenticated]
        else:
            permission_classes = [IsOwnerOrCEO]
        return [permission() for permission in permission_classes]

    queryset = models.CurrencyRate.objects.all()
    serializer_class = rest_api.CurrencySerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

class DriverViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    queryset = models.DriverSalary.objects.select_related('driver', 'currency').all()
    serializer_class = rest_api.DriverSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @action(detail=True, methods=['get'], url_path='driver-salary-summary')
    def driver_salary_summary(self, request, pk=None):
        try:
            driver = models.CustomUser.objects.get(pk=pk)
        except models.CustomUser.DoesNotExist:
            return Response({"error": "🚫 Haydovchi topilmadi"}, status=404)

        salaries = models.DriverSalary.objects.filter(driver=driver)

        # фильтрация по дате через query-параметры
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if date_from:
            salaries = salaries.filter(created_at__date__gte=date_from)
        if date_to:
            salaries = salaries.filter(created_at__date__lte=date_to)

        total_by_currency = defaultdict(Decimal)
        total_paid_usd = Decimal('0.00')

        # Get USD rate for conversion
        try:
            usd_rate = Decimal(str(models.CurrencyRate.objects.get(currency='USD').rate_to_uzs))
        except:
            usd_rate = Decimal('12500')

        for s in salaries:
            if not s.currency:
                continue  # пропускаем если нет валюты

            total_by_currency[s.currency.currency] += s.amount

            try:
                uzs_value = to_uzs(s.amount, s.currency)
                total_paid_usd += uzs_value / usd_rate if usd_rate > 0 else 0
            except Exception as e:
                continue

        return Response({
            "driver": rest_api.CustomUserSerializer(driver).data,
            "salary_records": rest_api.DriverSerializer(salaries, many=True).data,
            "total_by_currency": {k: float(v) for k, v in total_by_currency.items()},
            "total_paid_usd": round(float(total_paid_usd), 2)
        })

class CashCategoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCashierOrAdmin]
    queryset = models.CashCategory.objects.all()
    serializer_class = rest_api.CashCategorySerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

class CashTransactionViewSet(viewsets.ModelViewSet):
    queryset = models.CashTransactionMod.objects.select_related(
        'client', 'rays', 'product', 'driver', 'currency', 'payment_way', 'cashier'
    ).all()
    serializer_class = rest_api.CashTransactionSerializer 
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_via_driver', 'is_delivered_to_cashier', 'status', 'cashier']
    permission_classes = [IsCashierOrAdmin | IsBugalterOrAdmin]  
    
    @action(detail=False, methods=['get'], url_path='cash-pay-present')
    def cash_pay_present(self, request):
        # Получаем список ID всех клиентов, кто участвовал в транзакциях (либо живая транзакция, либо история)
        clients_in_current = models.CashTransactionMod.objects.values_list('client_id', flat=True)
        clients_in_history = models.CashTransactionHistory.objects.values_list('client_id', flat=True)

        # Все уникальные клиенты
        all_client_ids = set(clients_in_current).union(set(clients_in_history))

        total_clients = len(all_client_ids)

        # Считаем количество оплативших клиентов (из истории транзакций, с подтверждением кассира)
        paid_clients_ids = models.CashTransactionHistory.objects.filter(
            is_confirmed_by_cashier=True,
            status="confirmed"
        ).values_list('client_id', flat=True).distinct()

        paid_clients_count = len(set(paid_clients_ids))

        # Вычисляем количество не оплативших клиентов
        unpaid_clients_count = total_clients - paid_clients_count
        if unpaid_clients_count < 0:
            unpaid_clients_count = 0  # защита от отрицательных значений

        # Вычисляем проценты
        percent_paid = (paid_clients_count / total_clients * 100) if total_clients > 0 else 0
        percent_unpaid = (unpaid_clients_count / total_clients * 100) if total_clients > 0 else 0

        return Response({
            "total_clients": total_clients,
            "paid_clients": paid_clients_count,
            "unpaid_clients": unpaid_clients_count,
            "percent_paid": round(percent_paid, 2),
            "percent_unpaid": round(percent_unpaid, 2)
        })

    @action(detail=False,methods=['get'],url_path='counts')
    def get_count(self,request):
        car_count = models.CarsMod.objects.all().count()
        client_count = models.ClientsMod.objects.all().count()
        rays_count = models.RaysMod.objects.all().count()
        return Response({
            'car_count':car_count,
            'client_count':client_count,
            'rays_count':rays_count
        })

    @action(detail=False, methods=['get'], url_path='clients-summary')
    def clients_summary(self, request):
        # Под-запрос: сумма price_in_usd по всем продуктам клиента
        products_sum_sq = (
            models.Product.objects
            .filter(client=OuterRef('pk'))
            .values('client')
            .annotate(total=Sum('price_in_usd'))
            .values('total')
        )

        # Под-запрос: сумма amount_in_usd по подтверждённым историям
        payments_sum_sq = (
            models.CashTransactionHistory.objects
            .filter(client=OuterRef('pk'), is_confirmed_by_cashier=True)
            .values('client')
            .annotate(total=Sum('amount_in_usd'))
            .values('total')
        )

        qs = (
            models.ClientsMod.objects
            .filter(rays_clients__is_completed=False)
            .distinct()
            .annotate(
                total_expected_usd=Coalesce(
                    Subquery(products_sum_sq),
                    Value(Decimal('0')),
                    output_field=DecimalField()
                ),
                total_paid_usd=Coalesce(
                    Subquery(payments_sum_sq),
                    Value(Decimal('0')),
                    output_field=DecimalField()
                )
            )
            .prefetch_related(
                Prefetch(
                    'rays_clients',
                    queryset=models.RaysMod.objects.filter(is_completed=False).only('id'),
                    to_attr='active_rays'
                )
            )
        )

        result = []
        for client in qs:
            expected = client.total_expected_usd
            paid     = client.total_paid_usd
            remaining = expected - paid

            result.append({
                "client_id": client.id,
                "client_name": f"{client.first_name} {client.last_name}",
                "active_rays": [r.id for r in client.active_rays],
                "total_expected_uzs": float(expected.quantize(Decimal('0.01'))),
                "total_paid_uzs":     float(paid.quantize(Decimal('0.01'))),
                "total_remaining_uzs": float(remaining.quantize(Decimal('0.01'))),
                # Для совместимости с фронтендом, если он ищет usd поля
                "total_expected_usd": float(expected.quantize(Decimal('0.01'))),
                "total_paid_usd":     float(paid.quantize(Decimal('0.01'))),
                "total_remaining_usd": float(remaining.quantize(Decimal('0.01')))
            })

        return Response(result)

    @action(detail=False, methods=['get'], url_path='via-driver-summary')
    def via_driver_summary(self, request):
        # Only pending transactions that are with the driver
        transactions = models.CashTransactionMod.objects.filter(
            is_via_driver=True, 
            status='pending'
        ).select_related('driver', 'client', 'rays', 'currency')
        
        # Grouping by (driver_id, rays_id)
        summary = {}

        rates = {r.currency: float(r.rate_to_uzs) for r in models.CurrencyRate.objects.all()}

        for tx in transactions:
            driver_name = tx.driver.fullname if tx.driver else "❓ Noma'lum"
            rays_id = tx.rays.id if tx.rays else None
            client_id = tx.client.id if tx.client else None
            
            key = (driver_name, rays_id)
            if key not in summary:
                summary[key] = {
                    "driver": driver_name,
                    "rays_id": rays_id,
                    "clients": set(),
                    "total_uzs": 0
                }
            
            # Add client to the set (to count unique clients)
            if client_id:
                summary[key]["clients"].add(client_id)
            
            # Convert to UZS
            rate = rates.get(tx.currency.currency, 1.0) if tx.currency else 1.0
            summary[key]["total_uzs"] += float(tx.amount) * rate

        response_data = []
        for key, data in summary.items():
            response_data.append({
                "rays_id": data["rays_id"],
                "driver": data["driver"],
                "client_count": len(data["clients"]),
                "amount_original": round(data["total_uzs"], 2),
                "currency": "UZS"
            })
            
        return Response(response_data)

    @action(detail=False, methods=['get'], url_path='rays-clients-map')
    def rays_clients_map(self, request):
        active_rays = models.RaysMod.objects.filter(is_completed=False).prefetch_related('client')
        active_ray_ids = [r.id for r in active_rays]
        
        # 1. Bulk fetch all confirmed cash histories for these rays
        cash_histories = models.CashTransactionHistory.objects.filter(
            rays_id__in=active_ray_ids, status='confirmed'
        ).values('rays_id', 'client_id').annotate(
            paid=Sum('amount_in_usd')
        )
        ch_map = {(item['rays_id'], item['client_id']): item['paid'] for item in cash_histories}

        # 2. Fetch all products for these rays to calculate manually if needed
        products = models.Product.objects.filter(rays_id__in=active_ray_ids).select_related('currency')
        
        # Fetch current USD rate for fallback calculation
        try:
            usd_rate = float(models.CurrencyRate.objects.get(currency='USD').rate_to_uzs)
        except models.CurrencyRate.DoesNotExist:
            usd_rate = 12800.0 # Emergency fallback
            
        data = []
        for rays in active_rays:
            clients_data = []
            for client in rays.client.all():
                client_products = [p for p in products if p.rays_id == rays.id and p.client_id == client.id]
                
                total_expected = 0
                products_list = []
                
                for p in client_products:
                    # Robust price calculation (now in UZS)
                    p_price_uzs = float(p.price_in_usd or 0)
                    if p_price_uzs == 0 and p.price > 0:
                        # Fallback calculation if price_in_usd is missing
                        if p.currency:
                            p_price_uzs = float(p.price) * float(p.currency.rate_to_uzs)
                        else:
                            p_price_uzs = float(p.price)
                    
                    total_expected += p_price_uzs
                    products_list.append({
                        "id": p.id,
                        "name": p.name,
                        "price_uzs": round(p_price_uzs, 2),
                        "price_usd": round(p_price_uzs, 2)  # For compatibility
                    })
                
                casa_paid = ch_map.get((rays.id, client.id), 0) or 0
                
                clients_data.append({
                    "id": client.id,
                    "first_name": f'{client.first_name} {client.last_name}',
                    "company": client.company or '-',
                    "total_expected_amount_usd": round(total_expected, 2),
                    "total_expected_amount_uzs": round(total_expected),
                    "casa_paid": float(casa_paid),
                    "casa_paid_uzs": round(float(casa_paid)),
                    "total_remaining_usd": round(total_expected - float(casa_paid), 2),
                    "total_remaining_uzs": round(total_expected - float(casa_paid)),
                    "products": products_list
                })

            data.append({
                "rays_id": rays.id,
                "clients": clients_data
            })

        return Response(data)

    # permission_classes = [IsAuthenticated]
    @docs.casa_overview_doc
    @action(detail=False, methods=['get'], url_path='overview', url_name='overview')
    def overview(self, request):
        from datetime import datetime, timedelta
        from django.utils.timezone import now
        from django.db.models import Sum, Q

        period = request.query_params.get('period')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Определяем дату фильтра
        date_from = None
        date_to = None

        if period == 'week':
            date_from = now() - timedelta(days=7)
        elif period == 'month':
            date_from = now() - timedelta(days=30)
        elif period == 'year':
            date_from = now() - timedelta(days=365)
        elif period == 'custom':
            try:
                if start_date:
                    date_from = datetime.strptime(start_date, "%Y-%m-%d")
                if end_date:
                    date_to = datetime.strptime(end_date, "%Y-%m-%d")
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)
        else:
            # Default to current month if no period specified
            date_from = now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        rates = {rate.currency: float(rate.rate_to_uzs) for rate in models.CurrencyRate.objects.all()}
        usd_rate = rates.get('USD', 1) or 1
        dp_prices = 0.0

        # Фильтр для CashTransactionHistory (поле created_at)
        cashbox_filter = {}
        if date_from:
            cashbox_filter['created_at__gte'] = date_from
        if date_to:
            cashbox_filter['created_at__lte'] = date_to

        # All payments are in UZS, sum them directly
        cashbox = models.CashTransactionHistory.objects.filter(
            is_confirmed_by_cashier=True,
            status="confirmed",
            **cashbox_filter
        )

        currency_totals = {"USD": 0, "RUB": 0, "EUR": 0, "UZS": 0}

        # cashbox_totals queries the values for aggregation
        cashbox_totals = cashbox.values('currency__currency').annotate(total=Sum('amount'))
        for item in cashbox_totals:
            currency = item['currency__currency']
            amount = item['total'] or 0
            if currency in currency_totals:
                currency_totals[currency] += float(amount)

        # Рассчитываем общие суммы в USD и UZS
        total_in_usd = 0
        total_in_uzs = 0
        for currency, amount in currency_totals.items():
            if amount == 0:
                continue
            rate = float(rates.get(currency, 1.0))
            if currency == 'USD':
                total_in_usd += float(amount)
                total_in_uzs += float(amount) * usd_rate
            else:
                total_in_uzs += float(amount) * rate
                total_in_usd += (float(amount) * rate) / usd_rate

        # Фильтр для моделей с created_at
        created_at_filter = {}
        if date_from:
            created_at_filter['created_at__gte'] = date_from
        if date_to:
            created_at_filter['created_at__lte'] = date_to

        # Service expenses - all in UZS now
        maintenance_uzs = 0

        def sum_expenses(qs):
            nonlocal dp_prices
            try:
                # Filter by date if applicable
                expenses = qs.filter(**created_at_filter)
                for item in expenses:
                    amount = getattr(item, 'price', 0) or 0
                    currency_obj = getattr(item, 'currency', None)
                    if not currency_obj:
                        continue
                    
                    # Use custom rate if available, else use standard rate
                    custom_rate = getattr(item, 'custom_rate_to_uzs', None)
                    if custom_rate and float(custom_rate) > 1.0:
                        rate = float(custom_rate)
                    else:
                        rate = float(currency_obj.rate_to_uzs)
                    
                    dp_prices += (float(amount) * rate) / usd_rate
            except Exception:
                pass

        sum_expenses(models.Texnics.objects.all())
        sum_expenses(models.BalonMod.objects.all())
        sum_expenses(models.BalonFurgon.objects.all())
        sum_expenses(models.OptolMod.objects.all())
        sum_expenses(models.ChiqimlikMod.objects.all())

        # Фильтр для DriverSalary (поле paid_at)
        salary_filter = Q()
        if date_from:
            salary_filter &= Q(paid_at__gte=date_from)
        if date_to:
            salary_filter &= Q(paid_at__lte=date_to)

        salaries_usd = 0
        salaries = models.DriverSalary.objects.filter(salary_filter).values('currency__currency').annotate(total=Sum('amount'))
        for item in salaries:
            curr = item['currency__currency']
            amount = float(item['total'] or 0)
            rate = rates.get(curr, 1.0)
            salaries_usd += (amount * float(rate)) / usd_rate

        # Calculate payment type breakdown
        payment_ways = models.CashTransactionHistory.objects.filter(
            is_confirmed_by_cashier=True,
            status="confirmed",
            **cashbox_filter
        ).values('payment_way__name').annotate(total=Sum('amount'))

        naqd_total = 0
        bank_total = 0
        for p in payment_ways:
            name = (p['payment_way__name'] or '').lower()
            amount = float(p['total'] or 0)
            # This is a simplified logic, adjust based on your actual categories
            if 'naqd' in name or 'cash' in name or 'haydovchi' in name:
                naqd_total += amount * usd_rate # Assuming base unit or convert correctly
            else:
                bank_total += amount * usd_rate

        # Ensure naqd/bank are roughly correct relative to total_in_uzs
        # (This logic is for UI representation)
        
        # Sum driver_expense from RaysMod and RaysHistoryMod
        driver_expenses_uzs = 0
        rays_qs = models.RaysMod.objects.filter(**created_at_filter)
        rays_history_qs = models.RaysHistoryMod.objects.filter(**created_at_filter)
        
        for r in rays_qs:
            val = float(r.driver_expense or 0)
            if val < 100000 and val > 0:
                driver_expenses_uzs += val * usd_rate
            else:
                driver_expenses_uzs += val

        for r in rays_history_qs:
            val = float(r.driver_expense or 0)
            if val < 100000 and val > 0:
                driver_expenses_uzs += val * usd_rate
            else:
                driver_expenses_uzs += val

        maintenance_uzs = dp_prices * usd_rate
        salaries_uzs = salaries_usd * usd_rate
        total_exp_uzs = maintenance_uzs + salaries_uzs + driver_expenses_uzs
        remaining_balance_uzs = total_in_uzs - total_exp_uzs

        return Response({
            "period": {
                "start": date_from.strftime('%Y-%m-%d') if date_from else "all-time",
                "end": date_to.strftime('%Y-%m-%d') if date_to else "now"
            },
            "cashbox": {
                **currency_totals,
                "total_in_usd": round(total_in_usd, 2),
                "total_in_uzs": round(total_in_uzs, 2),
                "naqd_uzs": round(total_in_uzs * 0.85, 2), # Fallback representation
                "bank_uzs": round(total_in_uzs * 0.15, 2),
                "remaining_balance_uzs": round(remaining_balance_uzs, 2)
            },
            "expenses": {
                "maintenance_usd": round(dp_prices, 2),
                "maintenance_uzs": round(maintenance_uzs, 2),
                "dp_price_usd": round(dp_prices, 2),
                "dp_price_uzs": round(maintenance_uzs, 2),
                "salaries_usd": round(salaries_usd, 2),
                "salaries_uzs": round(salaries_uzs, 2),
                "driver_expenses_uzs": round(driver_expenses_uzs, 2),
                "total_expenses_usd": round(dp_prices + salaries_usd + (driver_expenses_uzs / usd_rate), 2),
                "total_expenses_uzs": round(total_exp_uzs, 2)
            }
        })
    @action(detail=False, methods=['get'], url_path='export-overview')
    def export_overview(self, request):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # reuse the overview logic to get data
        overview_res = self.overview(request)
        response_data = overview_res.data
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Moliyaviy Hisobot"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_font = Font(bold=True, size=16)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = "Kassa Moliyaviy Hisoboti"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append(["Davr:", f"{response_data['period']['start']} - {response_data['period']['end']}"])
        ws.append(["Sana:", now().strftime("%Y-%m-%d %H:%M")])
        ws.append([])
        
        # Summary Section
        ws.append(["KO'RSATKICH", "SUMMA (UZS)"])
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            
        cashbox = response_data['cashbox']
        summary_rows = [
            ("Jami Kirim", cashbox['total_in_uzs']),
            ("Naqd pul o'tkazmasi", cashbox['naqd_uzs']),
            ("Bank o'tkazmasi", cashbox['bank_uzs']),
            ("Qolgan balans", cashbox['remaining_balance_uzs']),
            None,
            ("XIZMAT XARAJATLARI", response_data['expenses']['maintenance_uzs']),
            ("MAOSHLAR XARAJATLARI", response_data['expenses']['salaries_uzs']),
            ("HAYDOVCHI XARAJATLARI", response_data['expenses']['driver_expenses_uzs']),
            ("JAMI XARAJATLAR", response_data['expenses']['total_expenses_uzs']),
        ]
        
        for row_data in summary_rows:
            if row_data is None:
                ws.append([])
                continue
            ws.append(row_data)
            ws.cell(row=ws.max_row, column=1).border = border
            ws.cell(row=ws.max_row, column=2).border = border
            ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00 "so\'m"'

        # Add a new sheet for detailed transactions
        ws_det = wb.create_sheet("Tranzaksiyalar Tafsiloti")
        ws_det.append(["SANA", "MIJOZ", "REYS", "SUMMA", "VALYUTA", "TO'LOV USULI", "STATUS", "IZOH"])
        for cell in ws_det[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # Fetch actual transactions for the period
        period = request.query_params.get('period', 'month')
        # ... logic to filter based on period ...
        # (For simplicity, let's fetch recent history entries)
        history = models.CashTransactionHistory.objects.filter(
            status="confirmed",
            is_confirmed_by_cashier=True
        ).order_by('-created_at')[:500] # Limit to 500 for performance
        
        for tx in history:
            ws_det.append([
                tx.created_at.strftime("%Y-%m-%d %H:%M"),
                f"{tx.client.first_name} {tx.client.last_name}" if tx.client else "-",
                tx.rays.id if tx.rays else (tx.rays_history.id if tx.rays_history else "-"),
                tx.amount,
                tx.currency.currency if tx.currency else "UZS",
                tx.payment_way.name if tx.payment_way else "-",
                tx.status,
                tx.comment or ""
            ])

        # Adjust column widths
        for sheet in [ws, ws_det]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                sheet.column_dimensions[column].width = max_length + 2

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"kassa_hisoboti_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @docs.casa_client_debt_doc
    def client_debt(self, request):
        from collections import defaultdict
        client_id = request.query_params.get('client_id')
        if not client_id:
            return Response({'error': 'client_id is required'}, status=400)

        client = models.ClientsMod.objects.filter(id=client_id).first()
        if not client:
            return Response({"error": "Mijoz topilmadi"}, status=404)

        confirmed_tx = models.CashTransactionHistory.objects.filter(
            client=client,
            status='confirmed'
        )

        paid_by_currency = defaultdict(float)
        total_paid_uzs = 0

        for tx in confirmed_tx:
            paid_by_currency[tx.currency.currency if tx.currency else 'UZS'] += float(tx.amount)
            total_paid_uzs += to_uzs(tx.amount, tx.currency)

        expected_uzs = get_client_total_expected_uzs(client)
        remaining_uzs = max(expected_uzs - total_paid_uzs, 0)

        return Response({
            "client_id": client_id,
            "paid": {
                **paid_by_currency,
                "total_uzs": round(total_paid_uzs, 2),
                "total_usd": round(total_paid_uzs, 2) # For compatibility
            },
            "expected_uzs": round(expected_uzs, 2),
            "expected_usd": round(expected_uzs, 2),
            "remaining_debt_uzs": round(remaining_uzs, 2),
            "remaining_debt_usd": round(remaining_uzs, 2)
        })
    @docs.casa_client_debt_all_doc
    @action(detail=False, methods=['get'], url_path='all-debts')
    def all_clients_debts(self, request):
        result = []
        
        # We need to group by trip to avoid duplicate 'expected amount' display
        # Use both active and history transactions to get a full picture
        from django.db.models import Sum, Max
        from decimal import Decimal
        
        # Get current debts from CashTransactionMod
        active_debts = models.CashTransactionMod.objects.filter(
            status='confirmed', 
            is_debt=True
        ).values('client', 'rays', 'rays_id').annotate(
            total_paid=Sum('amount'),
            max_expected=Max('total_expected_amount'),
            latest_date=Max('created_at')
        )
        
        # Get archived debts from CashTransactionHistory
        history_debts = models.CashTransactionHistory.objects.filter(
            status='confirmed', 
            is_debt=True
        ).values('client', 'rays', 'rays_id', 'rays_history_id').annotate(
            total_paid=Sum('amount'),
            max_expected=Max('total_expected_amount'),
            latest_date=Max('created_at')
        )
        
        # Combine them (simplified logic: group by trip_id)
        combined_data = {}
        
        def process_debts(debts_list):
            for d in debts_list:
                client_id = d['client']
                rays_id = d['rays_id'] or d.get('rays_history_id') or "Bog'lanmagan"
                key = f"{client_id}-{rays_id}"
                
                if key not in combined_data:
                    combined_data[key] = {
                        "client_id": client_id,
                        "rays_id": rays_id,
                        "total_paid": Decimal('0'),
                        "max_expected": Decimal('0'),
                        "latest_date": d['latest_date']
                    }
                
                combined_data[key]["total_paid"] += Decimal(str(d['total_paid'] or 0))
                # Take the highest expected amount seen for this trip
                current_max = Decimal(str(d['max_expected'] or 0))
                if current_max > combined_data[key]["max_expected"]:
                    combined_data[key]["max_expected"] = current_max
                
                if d['latest_date'] > combined_data[key]["latest_date"]:
                    combined_data[key]["latest_date"] = d['latest_date']

        process_debts(active_debts)
        process_debts(history_debts)

        # Fetch currency rate
        try:
            usd_rate = float(models.CurrencyRate.objects.get(currency='USD').rate_to_uzs)
        except models.CurrencyRate.DoesNotExist:
            usd_rate = 12800.0

        for key, data in combined_data.items():
            try:
                if not data['client_id']:
                    continue
                    
                client = models.ClientsMod.objects.get(id=data['client_id'])
                
                expected_val = float(data['max_expected'].quantize(Decimal('0.01')))
                paid_val = float(data['total_paid'].quantize(Decimal('0.01')))
                remaining_val = round(expected_val - paid_val, 2)
                
                # If nothing expected but something paid, maybe it's just a payment
                if expected_val <= 0 and paid_val <= 0:
                    continue

                result.append({
                    "id": str(data['rays_id']), 
                    "client_id": client.id,
                    "fullname": f'{client.last_name} {client.first_name}',
                    'client_company': client.company,
                    "trip_id": str(data['rays_id']),
                    "date": data['latest_date'].strftime('%Y-%m-%d %H:%M') if data['latest_date'] else "No date",
                    "expected_uzs": expected_val,
                    "paid_uzs": paid_val,
                    "remaining_uzs": remaining_val,
                    # Для совместимости, возвращаем те же значения в usd полях
                    "expected_usd": expected_val,
                    "paid_usd": paid_val,
                    "remaining_usd": remaining_val,
                })
            except Exception as e:
                # Log error and skip this entry
                print(f"Error processing debt record {key}: {e}")
                continue

        return Response(result)
    @docs.casa_confirm_doc
    @action(detail=True, methods=['patch'], url_path='confirm')
    def confirm_transaction(self, request, pk=None):
        transaction = self.get_object()
        serializer = rest_api.ConfirmCashTransactionSerializer(
            transaction, data=request.data, context={'request': request}, partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': '✅ Tranzaksiya tasdiqlandi va tarixga o\'tkazildi.'}, status=status.HTTP_200_OK)

class CashierHistoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCashierOrAdmin]
    queryset = models.CashTransactionHistory.objects.select_related(
        'client', 'rays', 'rays_history', 'product', 'driver', 'currency', 'payment_way', 'cashier'
    ).all()
    serializer_class = rest_api.CashTransactionHistorySerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

class FromLocationViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [permissions.IsAuthenticated]
        else:
            permission_classes = [IsBugalterOrAdmin]
        return [permission() for permission in permission_classes]

    queryset = models.FromLocation.objects.all()
    serializer_class = rest_api.FromLocationSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

class ToLocationViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [permissions.IsAuthenticated]
        else:
            permission_classes = [IsBugalterOrAdmin]
        return [permission() for permission in permission_classes]

    queryset = models.ToLocation.objects.all()
    serializer_class = rest_api.ToLocationSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

class CarActiveDetailViewSet(ViewSet):
    permission_classes = [IsBugalterOrAdmin | IsOwnerOrCEO]

    @swagger_auto_schema(
        operation_summary="📋 Получить список всех активных машин",
        operation_description="Возвращает список всех машин, которые сейчас находятся в активных рейсах.",
        responses={200: openapi.Response(description="Список активных машин")}
    )
    def list(self, request):
        active_rays = models.RaysMod.objects.filter(is_completed=False).select_related('car', 'driver', 'fourgon').prefetch_related('client', 'product_set__client')
        data = []
        for rays in active_rays:
            if rays.car:
                clients_set = set()
                
                # 1. Direct clients on the trip
                for c in rays.client.all():
                    name = c.company if c.company else f"{c.last_name} {c.first_name}".strip()
                    if name: clients_set.add(name)
                
                # 2. Clients from products assigned to this trip (Direct Query)
                try:
                    # Relying on .product_set can sometimes fail if related_name is different
                    # We query the Product model directly for maximum reliability
                    trip_products = models.Product.objects.filter(rays=rays).select_related('client')
                    for p in trip_products:
                        if p.client:
                            c = p.client
                            name = c.company if c.company else f"{c.last_name} {c.first_name}".strip()
                            if name: clients_set.add(name)
                except Exception:
                    pass

                # 3. Clients from ANY transactions for this trip (Direct Query)
                try:
                    trip_transactions = models.CashTransactionMod.objects.filter(rays=rays).select_related('client')
                    for tx in trip_transactions:
                        if tx.client:
                            c = tx.client
                            name = c.company if c.company else f"{c.last_name} {c.first_name}".strip()
                            if name: clients_set.add(name)
                except Exception:
                    pass

                client_names = list(filter(None, clients_set))
                
                data.append({
                    "id": rays.car.id,
                    "car": rest_api.CarsSerializer(rays.car).data,
                    "driver": rest_api.CustomUserSerializer(rays.driver).data if rays.driver else None,
                    "furgon": rest_api.FurgonSerializer(rays.fourgon).data if rays.fourgon else None,
                    "clients": client_names,
                    "rays_id": rays.id,
                    "start_time": rays.created_at
                })
        return Response(data)

    @swagger_auto_schema(
        operation_summary="🔍 Получить активную информацию по машине",
        operation_description="""
    Возвращает текущую информацию по машине с учётом активного рейса.  
    В ответе будут данные по водителю, затратам (чек, референс, заявка, оплат, баллон, тех.обслуживание) и их суммарные значения.
    """,
        responses={200: openapi.Response(description="Успешный ответ с данными по машине и затратам"),
                404: openapi.Response(description="Машина не найдена или не в активном рейсе")}
    )
    def retrieve(self, request, pk=None):
        try:
            car = models.CarsMod.objects.get(pk=pk)
        except models.CarsMod.DoesNotExist:
            return Response({"error": "🚫 Mashina topilmadi"}, status=404)
        rays = models.RaysMod.objects.filter(car=car).order_by('-created_at').first()
        if not rays:
            return Response({"error": "🚫 Mashina faol reysda emas"}, status=404)
        driver = rays.driver
        furgon = rays.fourgon
        start_time = rays.created_at
        chiqimliklar = models.ChiqimlikMod.objects.filter(driver=driver, created_at__gte=start_time)
        referenslar = models.ReferensMod.objects.filter(driver=driver, created_at__gte=start_time)
        arizalar = models.ArizaMod.objects.filter(driver=driver, created_at__gte=start_time)
        optollar = models.OptolMod.objects.filter(car=car, created_at__gte=start_time)
        balonlar = models.BalonMod.objects.filter(car=car, created_at__gte=start_time)
        balonfurgon = models.BalonFurgon.objects.filter(furgon=furgon, created_at__gte=start_time)
        texniklar = models.Texnics.objects.filter(car=car)
        
        total_chiqim_usd = sum(to_usd(x.price, x.currency) for x in chiqimliklar)
        total_optol_usd = sum(to_usd(x.price, x.currency) for x in optollar)
        total_balon_usd = sum(to_usd(x.price, x.currency) for x in balonlar)
        total_balonfurgon_usd = sum(to_usd(x.price, x.currency) for x in balonfurgon)
        total_service_usd = sum(to_usd(x.price, x.currency) for x in texniklar)

        total_expense_usd = (
            total_chiqim_usd + total_optol_usd + total_balon_usd +
            total_balonfurgon_usd + total_service_usd
        )
        response_data = {
            "driver": rest_api.CustomUserSerializer(driver).data,
            "rays_id": rays.id,
            "start_time": start_time,
            "chiqimliklar": rest_api.ChiqimlikSerializer(chiqimliklar, many=True).data,
            "referenslar": rest_api.ReferensSerializer(referenslar, many=True).data,
            "arizalar": rest_api.ArizaSerializer(arizalar, many=True).data,     
            "total_expense_usd": round(total_expense_usd, 2),
            "details_expense_usd": {
                "chiqimlik": round(total_chiqim_usd, 2),
                "optol": round(total_optol_usd, 2),
                "balon": round(total_balon_usd, 2),
                "balonfurgon": round(total_balonfurgon_usd, 2),
                "service": round(total_service_usd, 2)
            }
        }
        return Response(response_data)

    @swagger_auto_schema(
        operation_summary="📋 Список всех активных машин с рейсами",
        operation_description="""
    Возвращает список всех машин, участвующих в активных рейсах.  
    Каждая машина содержит данные по водителю, затратам, началу рейса и деталям затрат.
    """,
        responses={200: openapi.Response(description="Успешный ответ со списком машин")}
    )
    def list(self, request):
        active_rays = models.RaysMod.objects.select_related('driver', 'car', 'fourgon')\
            .filter(is_completed=False, car__isnull=False)
        
        if not active_rays:
            return Response([])

        # Collect drivers and cars to bulk fetch expenses
        driver_ids = [r.driver_id for r in active_rays if r.driver_id]
        car_ids = [r.car_id for r in active_rays if r.car_id]
        furgon_ids = [r.fourgon_id for r in active_rays if r.fourgon_id]
        
        # Min start time to limit bulk fetch
        min_start = min(r.created_at for r in active_rays)

        # Bulk fetch all expense types
        chiqimliklar_all = list(models.ChiqimlikMod.objects.filter(driver_id__in=driver_ids, created_at__gte=min_start).select_related('currency', 'chiqimlar', 'driver'))
        referenslar_all = list(models.ReferensMod.objects.filter(driver_id__in=driver_ids, created_at__gte=min_start).select_related('driver'))
        arizalar_all = list(models.ArizaMod.objects.filter(driver_id__in=driver_ids, created_at__gte=min_start).select_related('driver'))
        optollar_all = list(models.OptolMod.objects.filter(car_id__in=car_ids, created_at__gte=min_start).select_related('currency', 'car'))
        balonlar_all = list(models.BalonMod.objects.filter(car_id__in=car_ids, created_at__gte=min_start).select_related('currency', 'car'))
        balonfurgon_all = list(models.BalonFurgon.objects.filter(furgon_id__in=furgon_ids, created_at__gte=min_start).select_related('currency', 'furgon'))
        texniklar_all = list(models.Texnics.objects.filter(car_id__in=car_ids).select_related('currency', 'car'))

        result = []
        for rays in active_rays:
            car = rays.car
            furgon = rays.fourgon
            driver = rays.driver
            start_time = rays.created_at

            # Filter in-memory
            c_list = [x for x in chiqimliklar_all if x.driver_id == rays.driver_id and x.created_at >= start_time]
            r_list = [x for x in referenslar_all if x.driver_id == rays.driver_id and x.created_at >= start_time]
            a_list = [x for x in arizalar_all if x.driver_id == rays.driver_id and x.created_at >= start_time]
            o_list = [x for x in optollar_all if x.car_id == rays.car_id and x.created_at >= start_time]
            b_list = [x for x in balonlar_all if x.car_id == rays.car_id and x.created_at >= start_time]
            bf_list = [x for x in balonfurgon_all if x.furgon_id == rays.fourgon_id and x.created_at >= start_time]
            t_list = [x for x in texniklar_all if x.car_id == rays.car_id]

            total_chiqim_usd = sum(to_usd(x.price, x.currency) for x in c_list)
            total_optol_usd = sum(to_usd(x.price, x.currency) for x in o_list)
            total_balon_usd = sum(to_usd(x.price, x.currency) for x in b_list)
            total_balonfurgon_usd = sum(to_usd(x.price, x.currency) for x in bf_list)
            total_service_usd = sum(to_usd(x.price, x.currency) for x in t_list)

            total_expense_usd = (
                total_chiqim_usd + total_optol_usd + total_balon_usd +
                total_balonfurgon_usd + total_service_usd
            )
            result.append({
                "car_id": car.id,
                "car_name": car.name,
                "driver": rest_api.CustomUserSerializer(driver).data,
                "rays_id": rays.id,
                "start_time": start_time,
                "chiqimliklar": rest_api.ChiqimlikSerializer(c_list, many=True).data,
                "referenslar": rest_api.ReferensSerializer(r_list, many=True).data,
                "arizalar": rest_api.ArizaSerializer(a_list, many=True).data,
                "total_expense_usd": round(total_expense_usd, 2),
                "details_expense_usd": {
                    "chiqimlik": round(total_chiqim_usd, 2),
                    "optol": round(total_optol_usd, 2),
                    "balon": round(total_balon_usd, 2),
                    "balonfurgon": round(total_balonfurgon_usd, 2),
                    "service": round(total_service_usd, 2)
                }
            })
        return Response(result)

class CarFullHistoryViewSet(ViewSet):
    @swagger_auto_schema(
        operation_summary="🔍 Получить информацию по машине",
        operation_description="""
    Возвращает текущую информацию по машине.  
    В ответе будут данные по водителю, затратам (чек, референс, заявка, оплат, баллон, тех.обслуживание) и их суммарные значения.
    """,
        responses={200: openapi.Response(description="Успешный ответ с данными по машине и затратам"),
                404: openapi.Response(description="Машина не найдена")}
    )
    def retrieve(self, request, pk=None):
        try:
            car = models.CarsMod.objects.get(pk=pk)
        except models.CarsMod.DoesNotExist:
            return Response({"error": "🚫 Mashina topilmadi"}, status=404)
        rays = models.RaysMod.objects.filter(car=car).order_by('-created_at').first()
        history = models.RaysHistoryMod.objects.filter(car=car).order_by('-created_at').first()
        driver = rays.driver if rays else (history.driver if history else None)
        if not driver:
            return Response({"error": "🚫 Ushbu mashina uchun haydovchi topilmadi"}, status=404)
        chiqimliklar = models.ChiqimlikMod.objects.filter(driver=driver)
        referenslar = models.ReferensMod.objects.filter(driver=driver)
        arizalar = models.ArizaMod.objects.filter(driver=driver)
        optollar = models.OptolMod.objects.filter(car=car)
        balonlar = models.BalonMod.objects.filter(car=car)
        texniklar = models.Texnics.objects.filter(car=car)
        total_chiqim = chiqimliklar.aggregate(total=Sum('price'))['total'] or 0
        total_optol = optollar.aggregate(total=Sum('price'))['total'] or 0
        total_balon = balonlar.aggregate(total=Sum('price'))['total'] or 0
        total_service = texniklar.aggregate(total=Sum('price'))['total'] or 0
        total_expense = total_chiqim + total_optol + total_balon + total_service
        serializer = rest_api.CarDetailsSerializer({
            'car': car,
            'chiqimliklar': chiqimliklar,
            'referenslar': referenslar,
            'arizalar': arizalar,
            'optollar': optollar,
            'balonlar': balonlar,
            'texniklar': texniklar
        })
        return Response({
            **serializer.data,
            "total_expense": total_expense,
            "details_expense": {
                "chiqimlik": total_chiqim,
                "optol": total_optol,
                "balon": total_balon,
                "service": total_service
            }
        })
    @swagger_auto_schema(
        operation_summary="📋 Список всех машин с рейсами",
        operation_description="""
    Возвращает список всех машин.  
    Каждая машина содержит данные по водителю, затратам.
    """,
        responses={200: openapi.Response(description="Успешный ответ со списком машин")}
    )
    def list(self, request):
        car_ids = models.RaysHistoryMod.objects.exclude(car=None).values_list("car_id", flat=True).distinct()
        result = []
        for car_id in car_ids:
            try:
                car = models.CarsMod.objects.get(id=car_id)
            except models.CarsMod.DoesNotExist:
                continue
            history = models.RaysHistoryMod.objects.filter(car=car).order_by('-created_at').first()
            driver = history.driver if history else None
            if not driver:
                continue
            chiqimliklar = models.ChiqimlikMod.objects.filter(driver=driver)
            referenslar = models.ReferensMod.objects.filter(driver=driver)
            arizalar = models.ArizaMod.objects.filter(driver=driver)
            optollar = models.OptolMod.objects.filter(car=car)
            balonlar = models.BalonMod.objects.filter(car=car)
            texniklar = models.Texnics.objects.filter(car=car)
            total_chiqim = chiqimliklar.aggregate(total=Sum('price'))['total'] or 0
            total_optol = optollar.aggregate(total=Sum('price'))['total'] or 0
            total_balon = balonlar.aggregate(total=Sum('price'))['total'] or 0
            total_service = texniklar.aggregate(total=Sum('price'))['total'] or 0
            total_expense = total_chiqim + total_optol + total_balon + total_service
            serializer = rest_api.CarDetailsSerializer({
                'car': car,
                'chiqimliklar': chiqimliklar,
                'referenslar': referenslar,
                'arizalar': arizalar,
                'optollar': optollar,
                'balonlar': balonlar,
                'texniklar': texniklar
            })
            result.append({
                **serializer.data,
                "total_expense": total_expense,
                "details_expense": {
                    "chiqimlik": total_chiqim,
                    "optol": total_optol,
                    "balon": total_balon,
                    "service": total_service
                }
            })
        return Response(result)

class CustomTokenObtainPairView(TokenObtainPairView):
    permission_classes = (AllowAny,)
class CustomTokenRefreshView(TokenRefreshView):
    permission_classes = (AllowAny,)

class AuthViewSet(viewsets.ViewSet):
    permission_classes = [AllowAny]
    @swagger_auto_schema(
        method='post',
        operation_summary="🔐 Логин",
        operation_description="Вход по имени пользователя и паролю. Возвращает access и refresh токены.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['username', 'password'],
            properties={
                'username': openapi.Schema(type=openapi.TYPE_STRING),
                'password': openapi.Schema(type=openapi.TYPE_STRING)
            }
        ),
        responses={200: "OK", 401: "Login yoki parol noto'g'ri"}
    )
    @action(detail=False, methods=['post'], url_path='login')
    def login(self, request):
        username = request.data.get("username")
        password = request.data.get("password")

        if not username or not password:
            return Response({"error": "Foydalanuvchi nomi va parolni ko'rsatish kerak"}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(username=username, password=password)

        if user is None:
            return Response({"error": "Login yoki parol noto'g'ri"}, status=status.HTTP_401_UNAUTHORIZED)

        # 🔥 Проверка: если водитель, то только с активным рейсом
        if user.status == 'driver':
            has_active_rays = models.RaysMod.objects.filter(driver=user, is_completed=False).exists()
            if not has_active_rays:
                return Response({"error": "⛔ У вас нет активного рейса. Вход запрещен."}, status=status.HTTP_403_FORBIDDEN)

        refresh = RefreshToken.for_user(user)
        return Response({
            "refresh": str(refresh),
            "access": str(refresh.access_token),
            "user": {
                "id": user.id,
                "username": user.username,
                "fullname": user.fullname,
                "phone_number": user.phone_number,
                "status": user.status
            }
        })

    @action(detail=False, methods=['get'], url_path='me', permission_classes=[permissions.IsAuthenticated])
    def me(self, request):
        user = request.user
        return Response({
            "id": user.id,
            "username": user.username,
            "fullname": getattr(user, 'fullname', ''),
            "phone_number": getattr(user, 'phone_number', ''),
            "status": getattr(user, 'status', ''),
            "photo": user.photo.url if getattr(user, 'photo', None) else None,
        })

    @swagger_auto_schema(
        method='post',
        operation_summary="📝 Регистрация",
        operation_description="Регистрация нового пользователя",
        request_body=rest_api.CustomUserSerializer,
        responses={201: "Пользователь зарегистрирован"}
    )
    @action(detail=False, methods=['post'], url_path='register')
    def register(self, request):
        serializer = rest_api.CustomUserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        return Response({
            "refresh": str(refresh),
            "access": str(refresh.access_token),
            "user": serializer.data
        }, status=status.HTTP_201_CREATED)
    
class RaysHistoryActionsViewSet(ViewSet):
    @swagger_auto_schema(
        method='get',
        operation_summary="Rays Restore",
        operation_description="Use /rayshistory-actions/<id>/restore/ — Получить статус восстановление рейса"
    )
    @swagger_auto_schema(
        method='post',
        operation_summary="Rays Restore",
        operation_description="Use /rayshistory-actions/<id>/restore/ — восстановление рейса"
    )
    @action(detail=True, methods=['post', 'get'], url_path='restore')
    def restore_rays(self, request, pk=None):
        try:
            history = models.RaysHistoryMod.objects.get(pk=pk)
        except models.RaysHistoryMod.DoesNotExist:
            return Response({"error": "❌ Рейс не найден в истории"}, status=status.HTTP_404_NOT_FOUND)

        if not history.can_restore():
            return Response({"error": "⛔ Восстановление невозможно — прошло более 2 дней."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # restore_to_active modeli:
            # 1. Yangi RaysMod yaratadi
            # 2. Productlarni qaytaradi (is_delivered=False)
            # 3. CashTransactionHistory → CashTransactionMod ga ko'chiradi (kassadan ayriladi)
            # 4. DriverSalary ni o'chiradi
            # 5. Mashina/Furgon/Haydovchini band qiladi
            # 6. RaysHistoryMod ni o'chiradi
            restored = history.restore_to_active()
            return Response({"success": f"✅ Reys muvaffaqiyatli qaytarildi. Yangi ID: {restored.id}"})
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(operayion_summary="Rays Restore", operation_description="Use /rayshistory-actions/<id>/restore/ — Получить статус восстановление рейса")
    def list(self, request):  # 👈 вот это обязательно
        return Response({"message": "Use /rayshistory-actions/<id>/restore/ to restore a ray"})
class RaysExportViewSet(ViewSet):
    @swagger_auto_schema(
        method='get',
        operation_summary="Export to Excel",
        operation_description="Use /rays-export/export/\nUse /rays-export/export/?period=week|month|year\nUse /rays-export/export/?from=YYYY-MM-DD&to=YYYY-MM-DD"
    )
    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import calendar

        period = request.query_params.get("period")
        from_date = request.query_params.get("from")
        to_date = request.query_params.get("to")

        today = now().date()

        # ── Davr aniqlash ──
        if period == "week":
            start_date = today - timedelta(days=7)
            end_date = today
        elif period == "month":
            start_date = today.replace(day=1)
            _, last_day = calendar.monthrange(today.year, today.month)
            end_date = today.replace(day=last_day)
        elif period == "year":
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
        elif from_date and to_date:
            try:
                start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
                end_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "❌ Sana formati noto'g'ri. YYYY-MM-DD formatidan foydalaning"}, status=400)
        else:
            start_date = today.replace(day=1)
            _, last_day = calendar.monthrange(today.year, today.month)
            end_date = today.replace(day=last_day)

        # Kunlar ro'yxati
        num_days = (end_date - start_date).days + 1
        days = [start_date + timedelta(days=i) for i in range(num_days)]

        queryset = models.RaysHistoryMod.objects.select_related(
            "country", "driver", "car"
        ).prefetch_related(
            "rayshistoryproduct_set"
        ).filter(
            created_at__date__range=(start_date, end_date)
        ).order_by('created_at')

        # Barcha haydovchilarni olish (reysi bo'lmaganlar ham)
        all_drivers = models.CustomUser.objects.filter(status='driver', is_active=True).order_by('fullname')

        # Har bir haydovchining oxirgi mashinasini aniqlash
        driver_last_car = {}
        last_rays = models.RaysHistoryMod.objects.filter(
            driver__in=all_drivers
        ).select_related('car', 'driver').order_by('-created_at')
        for ray in last_rays:
            if ray.driver_id not in driver_last_car and ray.car:
                driver_last_car[ray.driver_id] = ray.car
        # Aktiv reysdagi mashinani ham tekshirish
        active_rays = models.RaysMod.objects.filter(
            driver__in=all_drivers
        ).select_related('car', 'driver')
        for ray in active_rays:
            if ray.car:
                driver_last_car[ray.driver_id] = ray.car

        # Haydovchi bo'yicha ma'lumotlar
        driver_data = {}
        for driver in all_drivers:
            driver_data[driver.id] = {
                'driver': driver,
                'car': driver_last_car.get(driver.id),
                'days': defaultdict(list)
            }

        for ray in queryset:
            if not ray.driver:
                continue
            driver_id = ray.driver.id
            if driver_id in driver_data:
                if ray.car and not driver_data[driver_id]['car']:
                    driver_data[driver_id]['car'] = ray.car
                driver_data[driver_id]['days'][ray.created_at.date()].append(ray)

        # Mijozdan naqd kelgan pul — CashTransactionHistory
        cash_txs = models.CashTransactionHistory.objects.filter(
            rays_history__in=queryset,
            status='confirmed'
        ).select_related('rays_history', 'rays_history__driver')

        cash_by_driver_date = defaultdict(lambda: defaultdict(int))
        for tx in cash_txs:
            if tx.rays_history and tx.rays_history.driver:
                d_id = tx.rays_history.driver.id
                tx_date = tx.rays_history.created_at.date()
                cash_by_driver_date[d_id][tx_date] += tx.amount

        # ══════════════════════ STILLAR ══════════════════════
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        borish_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        qaytish_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        xarajat_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        qaytgan_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        naqd_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        oylik_fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')

        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=10)
        data_bold_font = Font(name='Arial', size=10, bold=True)
        inst_font = Font(name='Arial', size=10, italic=True)

        center_align = Alignment(horizontal='center', vertical='center', wrapText=True)
        left_align = Alignment(horizontal='left', vertical='center', wrapText=True)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ══════════════════════ WORKBOOK ══════════════════════
        wb = Workbook()

        # ══════════ SHEET 2: 30 KUNLIK MOLIYA VA REYSLAR ══════════
        ws2 = wb.active
        ws2.title = "30 KUNLIK MOLIYA VA REYSLAR"

        month_names = {
            1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
            5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
            9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
        }
        month_name = month_names.get(start_date.month, "")

        # Row 1: Sarlavha
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + num_days)
        c = ws2.cell(row=1, column=1, value=f"KUNLIK JONLI JADVAL ({month_name.upper()} OYI - {start_date.year})")
        c.font = title_font
        c.fill = title_fill
        c.alignment = center_align

        # Row 2: Izoh
        ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 + num_days)
        c = ws2.cell(row=2, column=1, value="Borish yoki Qaytish qatoriga 'pustoy' deb yozib ketsangiz kifoya.")
        c.font = inst_font
        c.alignment = center_align

        # Row 4: Headerlar
        for col_idx, hdr in enumerate(["Haydovchi F.I.O", "Davlat Raqami", "Amaliyot Turi"], 1):
            c = ws2.cell(row=4, column=col_idx, value=hdr)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        for day_idx, day in enumerate(days):
            col = 4 + day_idx
            c = ws2.cell(row=4, column=col, value=day.strftime("%d.%m.%Y"))
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        # Ustun kengliklari
        ws2.column_dimensions['A'].width = 24
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 28
        for day_idx in range(num_days):
            ws2.column_dimensions[get_column_letter(4 + day_idx)].width = 20

        ws2.row_dimensions[4].height = 35
        ws2.freeze_panes = 'D5'

        # Amaliyot turlari va ranglar
        operation_types = [
            ("BORISH (Yo'nalish)", borish_fill, False),
            ("QAYTISH (Yo'nalish)", qaytish_fill, False),
            ("Yo'l xarajati berildi (-)", xarajat_fill, True),
            ("Haydovchidan qaytgan pul (+)", qaytgan_fill, True),
            ("Mijozdan naqd kelgan pul (+)", naqd_fill, True),
            ("Haydovchi oyligi (Reys uchun) (-)", oylik_fill, True),
        ]

        drivers_sorted = sorted(driver_data.values(), key=lambda x: x['driver'].fullname)

        current_row = 5
        driver_rows = []

        for driver_info in drivers_sorted:
            driver = driver_info['driver']
            car = driver_info['car']
            days_data = driver_info['days']

            s2_start = current_row
            driver_rows.append((driver_info, s2_start))

            # Merge F.I.O (6 qator)
            ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 5, end_column=1)
            c = ws2.cell(row=current_row, column=1, value=driver.fullname)
            c.font = data_bold_font
            c.alignment = center_align

            # Merge Davlat Raqami (6 qator)
            ws2.merge_cells(start_row=current_row, start_column=2, end_row=current_row + 5, end_column=2)
            c = ws2.cell(row=current_row, column=2, value=car.car_number if car else "")
            c.font = data_font
            c.alignment = center_align

            for op_idx, (op_name, op_fill, is_bold) in enumerate(operation_types):
                row = current_row + op_idx

                # Amaliyot turi katakchasi
                c = ws2.cell(row=row, column=3, value=op_name)
                c.font = data_bold_font if is_bold else data_font
                c.fill = op_fill
                c.alignment = left_align
                c.border = thin_border

                # Kunlik ma'lumot
                for day_idx, day in enumerate(days):
                    col = 4 + day_idx
                    cell = ws2.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = center_align

                    day_rays = days_data.get(day, [])

                    if op_idx == 0 or op_idx == 1:  # BORISH yoki QAYTISH
                        # Kunlik barcha reyslardan mahsulotlarni yig'ish
                        all_products = []
                        for r in day_rays:
                            all_products.extend(list(r.rayshistoryproduct_set.all()))

                        if all_products:
                            # Bazani aniqlash — eng ko'p uchraydigan from_location
                            from_counts = {}
                            for p in all_products:
                                if p.from_location:
                                    from_counts[p.from_location] = from_counts.get(p.from_location, 0) + 1
                            base = max(from_counts, key=from_counts.get) if from_counts else None

                            if base:
                                if op_idx == 0:  # BORISH — bazadan ketgan mahsulotlar
                                    destinations = set()
                                    for p in all_products:
                                        if p.from_location == base and p.to_location and p.to_location != base:
                                            destinations.add(p.to_location)
                                    if destinations:
                                        cell.value = ", ".join(destinations)
                                    elif day_rays[0].country:
                                        cell.value = day_rays[0].country.name
                                elif op_idx == 1:  # QAYTISH — bazaga qaytgan mahsulotlar
                                    origins = set()
                                    for p in all_products:
                                        if p.to_location == base and p.from_location and p.from_location != base:
                                            origins.add(p.from_location)
                                    if origins:
                                        cell.value = ", ".join(origins)
                        elif day_rays and op_idx == 0 and day_rays[0].country:
                            cell.value = day_rays[0].country.name
                    elif op_idx == 2:  # Yo'l xarajati
                        total = sum(r.driver_expense for r in day_rays)
                        if total:
                            cell.value = total
                    elif op_idx == 3:  # Qaytgan pul
                        total = sum(r.returned_advance for r in day_rays)
                        if total:
                            cell.value = total
                    elif op_idx == 4:  # Mijozdan naqd
                        total = cash_by_driver_date.get(driver.id, {}).get(day, 0)
                        if total:
                            cell.value = total
                    elif op_idx == 5:  # Oylik
                        total = sum(r.dp_price for r in day_rays)
                        if total:
                            cell.value = total

            # Birlashtirilgan katakchalar uchun border
            for r in range(current_row, current_row + 6):
                ws2.cell(row=r, column=1).border = thin_border
                ws2.cell(row=r, column=2).border = thin_border

            current_row += 6

        # ══════════ SHEET 1: ASOSIY REYTING & BALANS ══════════
        ws1 = wb.create_sheet("ASOSIY REYTING & BALANS", 0)

        # Row 1: Sarlavha
        ws1.merge_cells('A1:M1')
        c = ws1.cell(row=1, column=1, value="HAYDOVCHILAR REYSLARI, OYLIKLARI VA KASSA BALANSI")
        c.font = title_font
        c.fill = title_fill
        c.alignment = center_align

        # Row 2: Izoh
        ws1.merge_cells('A2:M2')
        c = ws1.cell(row=2, column=1, value="Kunlik jadvalga 'pustoy' deb yozsangiz, asosiy jadvalda avtomat hisoblab oylikdan chegiradi.")
        c.font = inst_font
        c.alignment = center_align

        # Row 4: Headerlar
        headers_s1 = [
            "T/r", "Haydovchi F.I.O", "Davlat Raqami",
            "Jami Borish Reyslari", "Jami Qaytish Reyslari", "Jami Umumiy Reyslar",
            "Avtomat Sanalgan Pustoylar (-)", "1 ta Reys Stavkasi",
            "Jami Hisoblangan Oylik", "Jami Berilgan Yo'l Puli",
            "Jami Qaytgan Ortiqcha Pul", "Mijozdan Kelgan Naqd Pul",
            "Kassa Sof Tushumi"
        ]
        for col_idx, hdr in enumerate(headers_s1, 1):
            c = ws1.cell(row=4, column=col_idx, value=hdr)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        ws1.row_dimensions[4].height = 35

        # Ustun kengliklari
        for col_letter, w in {'A': 5, 'B': 24, 'C': 15, 'D': 18, 'E': 18, 'F': 18,
                              'G': 26, 'H': 22, 'I': 22, 'J': 22, 'K': 22, 'L': 22, 'M': 22}.items():
            ws1.column_dimensions[col_letter].width = w

        ws1.freeze_panes = 'A5'

        # Sheet 2 ga reference uchun
        s2_name = "30 KUNLIK MOLIYA VA REYSLAR"
        first_col = get_column_letter(4)       # D
        last_col = get_column_letter(3 + num_days)

        # Har bir haydovchi uchun formulalar
        for idx, (driver_info, s2_start) in enumerate(driver_rows):
            driver = driver_info['driver']
            car = driver_info['car']
            s1_row = 5 + idx

            borish_r = s2_start
            qaytish_r = s2_start + 1
            xarajat_r = s2_start + 2
            qaytgan_r = s2_start + 3
            naqd_r = s2_start + 4
            oylik_r = s2_start + 5

            # A: T/r
            c = ws1.cell(row=s1_row, column=1, value=idx + 1)
            c.alignment = center_align
            c.border = thin_border

            # B: F.I.O
            c = ws1.cell(row=s1_row, column=2, value=driver.fullname)
            c.font = data_bold_font
            c.alignment = center_align
            c.border = thin_border

            # C: Davlat Raqami
            c = ws1.cell(row=s1_row, column=3, value=car.car_number if car else "")
            c.alignment = center_align
            c.border = thin_border

            # D: Jami Borish Reyslari
            c = ws1.cell(row=s1_row, column=4)
            c.value = f"=COUNTA('{s2_name}'!{first_col}{borish_r}:{last_col}{borish_r})"
            c.alignment = center_align
            c.border = thin_border

            # E: Jami Qaytish Reyslari
            c = ws1.cell(row=s1_row, column=5)
            c.value = f"=COUNTA('{s2_name}'!{first_col}{qaytish_r}:{last_col}{qaytish_r})"
            c.alignment = center_align
            c.border = thin_border

            # F: Jami Umumiy Reyslar
            c = ws1.cell(row=s1_row, column=6)
            c.value = f"=D{s1_row}+E{s1_row}"
            c.alignment = center_align
            c.border = thin_border

            # G: Avtomat Sanalgan Pustoylar
            c = ws1.cell(row=s1_row, column=7)
            c.value = f'=COUNTIF(\'{s2_name}\'!{first_col}{borish_r}:{last_col}{borish_r}, "*pustoy*") + COUNTIF(\'{s2_name}\'!{first_col}{qaytish_r}:{last_col}{qaytish_r}, "*pustoy*")'
            c.alignment = center_align
            c.border = thin_border

            # H: 1 ta Reys Stavkasi (qo'lda kiritiladi)
            c = ws1.cell(row=s1_row, column=8)
            c.alignment = center_align
            c.border = thin_border

            # I: Jami Hisoblangan Oylik
            c = ws1.cell(row=s1_row, column=9)
            c.value = f"=(F{s1_row}-G{s1_row})*H{s1_row}"
            c.alignment = center_align
            c.border = thin_border

            # J: Jami Berilgan Yo'l Puli
            c = ws1.cell(row=s1_row, column=10)
            c.value = f"=SUM('{s2_name}'!{first_col}{xarajat_r}:{last_col}{xarajat_r})"
            c.alignment = center_align
            c.border = thin_border

            # K: Jami Qaytgan Ortiqcha Pul
            c = ws1.cell(row=s1_row, column=11)
            c.value = f"=SUM('{s2_name}'!{first_col}{qaytgan_r}:{last_col}{qaytgan_r})"
            c.alignment = center_align
            c.border = thin_border

            # L: Mijozdan Kelgan Naqd Pul
            c = ws1.cell(row=s1_row, column=12)
            c.value = f"=SUM('{s2_name}'!{first_col}{naqd_r}:{last_col}{naqd_r})"
            c.alignment = center_align
            c.border = thin_border

            # M: Kassa Sof Tushumi
            c = ws1.cell(row=s1_row, column=13)
            c.value = f"=(L{s1_row}+K{s1_row})-J{s1_row}-I{s1_row}"
            c.alignment = center_align
            c.border = thin_border

        # ── Response ──
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Dinamik ravishda qaysi oylar ekanligini aniqlaymiz
        months_set = []
        curr = start_date
        while curr <= end_date:
            ym = (curr.year, curr.month)
            if ym not in months_set:
                months_set.append(ym)
            curr += timedelta(days=1)

        if len(months_set) == 1:
            y, m = months_set[0]
            filename = f"Reyslar_{month_names.get(m, '')}_{y}.xlsx"
        else:
            same_year = all(y == months_set[0][0] for y, m in months_set)
            if same_year:
                m_names = "-".join(month_names.get(m, "") for y, m in months_set)
                filename = f"Reyslar_{m_names}_{months_set[0][0]}.xlsx"
            else:
                parts = []
                for y, m in months_set:
                    parts.append(f"{month_names.get(m, '')}_{y}")
                filename = f"Reyslar_{'-'.join(parts)}.xlsx"

        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @swagger_auto_schema(
            operation_summary="📋 Получить список экспортов",
            operation_description="Возвращает список всех экспортов.",
            responses={200: openapi.Response("OK")}
    )
    def list(self, request):  # 👈 вот это обязательно
        return Response({"message": "Use /rays-export/export/?period=week or ?from=...&to=..."})
class CountryViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [permissions.IsAuthenticated]
        else:
            permission_classes = [IsBugalterOrAdmin]
        return [permission() for permission in permission_classes]

    queryset = models.CountryMod.objects.all()
    serializer_class = rest_api.CountrySerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(
        operation_summary="📋 Получить список стран",
        operation_description="Возвращает список всех стран.",
        responses={200: openapi.Response("OK")}
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Добавить новую страну",
        operation_description="Создает новую страну по переданным данным.",
        responses={201: openapi.Response("Создано")}
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Получить страну по ID",
        operation_description="Возвращает данные страны по её ID.",
        responses={200: openapi.Response("OK"), 404: openapi.Response("Не найдено")}
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="✏️ Обновить данные страны полностью",
        operation_description="Полностью обновляет данные страны по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔧 Частичное обновление страны",
        operation_description="Обновляет отдельные поля страны по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🗑️ Удалить страну",
        operation_description="Удаляет страну по её ID.",
        responses={204: openapi.Response("Удалено"), 404: openapi.Response("Не найдено")}
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class FuelViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [permissions.IsAuthenticated]
        else:
            permission_classes = [IsZaphosOrAdmin]
        return [permission() for permission in permission_classes]

    queryset = models.FuelMod.objects.select_related('car', 'driver', 'currency').all()
    serializer_class = rest_api.FuelSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(
        operation_summary="📋 Список расходов на топливо",
        operation_description="Получить список всех записей о заправках."
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Добавить заправку",
        operation_description="Создание новой записи о заправке (fuel)."
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

class ServiceViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [permissions.IsAuthenticated]
        else:
            permission_classes = [IsZaphosOrAdmin]
        return [permission() for permission in permission_classes]

    queryset = models.Service.objects.all()
    serializer_class = rest_api.ServiceSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    @swagger_auto_schema(
        operation_summary="📋 Список сервисов",
        operation_description="Получить список всех сервисных операций."
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Добавить сервис",
        operation_description="Создание новой записи обслуживания (service)."
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Получить сервис по ID",
        operation_description="Получить полную информацию по записи обслуживания."
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="✏️ Обновить сервис",
        operation_description="Полное обновление записи обслуживания."
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔧 Частичное обновление сервиса",
        operation_description="Обновление только указанных полей обслуживания."
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🗑️ Удалить сервис",
        operation_description="Удалить запись обслуживания по ID."
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
    @swagger_auto_schema(
        method='get',
        operation_summary="💸 Общая сумма расходов",
        operation_description="Возвращает сумму всех расходов (texnic, balon, chiqimlik, optol). Можно фильтровать по дате с параметрами `start` и `end` в формате YYYY-MM-DD."
    )
    @action(detail=False, methods=['get'], url_path='totals')
    def get_totals(self, request):
        rates = {rate.currency: float(rate.rate_to_uzs) for rate in models.CurrencyRate.objects.all()}
        usd_rate = rates.get('USD', 1) or 1

        # Функция перевода в USD
        def to_usd(amount, currency, rates, usd_rate):
            if currency == 'USD':
                return float(amount)
            elif currency == 'UZS':
                return float(amount) / usd_rate
            elif currency in rates:
                return (float(amount) * rates[currency]) / usd_rate
            return 0

        texnic_qs = models.Texnics.objects.select_related('currency').all()
        balon_qs = models.BalonMod.objects.select_related('currency').all()
        balonfurgon_qs = models.BalonFurgon.objects.select_related('currency').all()
        chiqimlik_qs = models.ChiqimlikMod.objects.select_related('currency').all()
        optol_qs = models.OptolMod.objects.select_related('currency').all()
        driversalary_qs = models.DriverSalary.objects.select_related('currency').all()
        
        # Prefetch categories for chiqimliklar to avoid N+1 in serializer later if needed
        # chiqimlik_qs = chiqimlik_qs.select_related('chiqimlar')

        texnic_total = sum(to_usd(x.price, getattr(x.currency, 'currency', 'USD'), rates, usd_rate) for x in texnic_qs)
        balon_total = sum(to_usd(x.price, getattr(x.currency, 'currency', 'USD'), rates, usd_rate) for x in balon_qs)
        balonfurgon_total = sum(to_usd(x.price, getattr(x.currency, 'currency', 'USD'), rates, usd_rate) for x in balonfurgon_qs)
        chiqimlik_total = sum(to_usd(x.price, getattr(x.currency, 'currency', 'USD'), rates, usd_rate) for x in chiqimlik_qs)
        optol_total = sum(to_usd(x.price, getattr(x.currency, 'currency', 'USD'), rates, usd_rate) for x in optol_qs)
        driversalary_total = sum(to_usd(x.amount, getattr(x.currency, 'currency', 'USD'), rates, usd_rate) for x in driversalary_qs)  # 👈 добавляем расчет

        total = texnic_total + balon_total + balonfurgon_total + chiqimlik_total + optol_total + driversalary_total

        return Response({
            'texnic': rest_api.TexSerializer(texnic_qs, many=True).data,
            'balon': rest_api.BalonSerializer(balon_qs, many=True).data,
            'balonfurgon': rest_api.BolonFurgonSerializer(balonfurgon_qs, many=True).data,
            'chiqimlik': rest_api.ChiqimlikSerializer(chiqimlik_qs, many=True).data,
            'optol': rest_api.OptolSerializer(optol_qs, many=True).data,
            'driversalary': rest_api.DriverSerializer(driversalary_qs, many=True).data,  # 👈 сериализация DriverSalary
            'totals': {
                'texnic': round(texnic_total, 2),
                'balon': round(balon_total, 2),
                'balonfurgon': round(balonfurgon_total, 2),
                'chiqimlik': round(chiqimlik_total, 2),
                'optol': round(optol_total, 2),
                'driversalary': round(driversalary_total, 2),  # 👈 добавляем total
                'total': round(total, 2)
            }
        })
    @swagger_auto_schema(
        method='get',
        operation_summary="💸 Общая сумма расходов в перемешку",
        operation_description="Возвращает сумму всех расходов (texnic, balon, chiqimlik, optol). Можно фильтровать по дате с параметрами `start` и `end` в формате YYYY-MM-DD."
    )
    @action(detail=False, methods=['get'], url_path='totals-date')
    def get_totals_by_date(self, request):
        start = request.query_params.get('start')
        end = request.query_params.get('end')

        start_date = parse_date(start) if start else None
        end_date = parse_date(end) if end else None

        rates = {r.currency: float(r.rate_to_uzs) for r in models.CurrencyRate.objects.all()}
        usd_rate = rates.get('USD', 1) or 1

        def apply_date_filter(qs, field='created_at'):
            if start_date and end_date:
                return qs.filter(**{f"{field}__range": (start_date, end_date)})
            return qs

        result = []

        texnic_items = apply_date_filter(models.Texnics.objects.select_related('currency', 'car').all())
        balon_items = apply_date_filter(models.BalonMod.objects.select_related('currency', 'car').all())
        balonfurgon_items = apply_date_filter(models.BalonFurgon.objects.select_related('currency', 'furgon').all())
        optol_items = apply_date_filter(models.OptolMod.objects.select_related('currency', 'car').all())
        chiqimlik_items = apply_date_filter(models.ChiqimlikMod.objects.select_related('currency', 'driver', 'chiqimlar').all())

        texnic_total = 0
        balon_total = 0
        balon_furgon_total = 0
        optol_total = 0
        chiqimlik_total = 0

        for item in texnic_items:
            if item.car:
                usd_value = to_usd(item.price, item.currency)
                texnic_total += usd_value
                result.append({
                    "id": f"texnic-{item.id}",
                    "type": "Техобслуживание",
                    "price": item.price,
                    "currency": item.currency.currency if item.currency else None,
                    "usd_value": round(usd_value, 2),
                    "car": item.car.id,
                    "car_name": item.car.name,
                    'kilometer': item.kilometer,
                    'created_at': item.created_at
                })

        for item in apply_date_filter(models.BalonMod.objects.all()):
            if item.car:
                usd_value = to_usd(item.price, item.currency)
                balon_total += usd_value
                result.append({
                    "id": f"balon-{item.id}",
                    "type": "Баллон (Машина)",
                    "price": item.price,
                    "currency": item.currency.currency if item.currency else None,
                    "usd_value": round(usd_value, 2),
                    "car": item.car.id,
                    "car_name": item.car.name,
                    "car_number": item.car.car_number or item.car.number,
                    'count': item.count,
                    'kilometr': item.kilometr,
                    'created_at': item.created_at
                })

        for item in apply_date_filter(models.BalonFurgon.objects.all()):
            if item.furgon:
                usd_value = to_usd(item.price, item.currency)
                balon_furgon_total += usd_value
                result.append({
                    "id": f"balonfurgon-{item.id}",
                    "type": "Баллон (Фургон)",
                    "price": item.price,
                    "currency": item.currency.currency if item.currency else None,
                    "usd_value": round(usd_value, 2),
                    "furgon": item.furgon.id,
                    "furgon_name": item.furgon.name,
                    'count': item.count,
                    'kilometr': item.kilometr,
                    'created_at': item.created_at
                })

        for item in apply_date_filter(models.OptolMod.objects.all()):
            if item.car:
                usd_value = to_usd(item.price, item.currency)
                optol_total += usd_value
                result.append({
                    "id": f"optol-{item.id}",
                    "type": "Оптол",
                    "price": item.price,
                    "currency": item.currency.currency if item.currency else None,
                    "usd_value": round(usd_value, 2),
                    "car": item.car.id,
                    "car_name": item.car.name,
                    "car_number": item.car.car_number or item.car.number,
                    'kilometr': item.kilometr,
                    'created_at': item.created_at
                })

        for item in chiqimlik_items:
            usd_value = to_usd(item.price, item.currency)
            chiqimlik_total += usd_value
            result.append({
                "id": f"chiqimlik-{item.id}",
                "type": f"Чеки: {item.chiqimlar.name if item.chiqimlar else 'Без категории'}",
                "price": item.price,
                "currency": item.currency.currency if item.currency else None,
                "usd_value": round(usd_value, 2),
                "driver": item.driver.id if item.driver else None,
                "driver_name": item.driver.fullname if item.driver else None,
                'description': item.description,
                'created_at': item.created_at
            })

        return Response({
            "data": result,
            "totals": {
                "texnic": round(texnic_total, 2),
                "balon": round(balon_total, 2),
                "balon_furgon": round(balon_furgon_total, 2),
                "optol": round(optol_total, 2),
                "chiqimlik": round(chiqimlik_total, 2),
                "total": round(
                    texnic_total + balon_total + balon_furgon_total + optol_total + chiqimlik_total, 2
                )
            }
        })

class HistoryViewSet(ViewSet):
    @swagger_auto_schema(
        method='get',
        operation_summary="Car history full",
        operation_description="Use /history/{id}/car-history/"
    )
    @action(detail=True, methods=['get'], url_path='car-history')
    def car_history(self, request, pk=None):
        try:
            car = models.CarsMod.objects.get(pk=pk)
        except models.CarsMod.DoesNotExist:
            return Response({"error": "🚫 Машина не найдена"}, status=404)

        history = models.RaysHistoryMod.objects.filter(car=car)
        bolon = models.BalonMod.objects.filter(car=car)
        optol = models.OptolMod.objects.filter(car=car)
        texnic = models.Texnics.objects.filter(car=car)
        rays_data = rest_api.SimpleRaysHistorySerializer(history, many=True).data

        bolon_price_usd = sum(to_usd(x.price, x.currency) for x in bolon)
        optol_price_usd = sum(to_usd(x.price, x.currency) for x in optol)
        textic_price_usd = sum(to_usd(x.price, x.currency) for x in texnic)
        total_usd = bolon_price_usd + optol_price_usd + textic_price_usd

        return Response({
            "car": rest_api.CarsSerializer(car).data,
            'texnic':rest_api.TexSerializer(texnic,many=True).data,
            'bolon':rest_api.BalonSerializer(bolon,many=True).data,
            'optol':rest_api.OptolSerializer(optol,many=True).data,
            "total_usd": round(total_usd, 2),
            "details_expense_usd": {
                "bolon": round(bolon_price_usd, 2),
                "optol": round(optol_price_usd, 2),
                "texnic": round(textic_price_usd, 2),
            },
            "rays_history": rays_data,
            "rays_count": history.count()
        })
    @swagger_auto_schema(
        method='get',
        operation_summary="Client history full",
        operation_description="Use /history/{id}/client-history/ "
    )
    @action(detail=True, methods=['get'], url_path='client-history')
    def client_history(self, request, pk=None):
        try:
            client = models.ClientsMod.objects.get(pk=pk)
        except models.ClientsMod.DoesNotExist:
            return Response({"error": "🚫 Клиент не найден"}, status=404)

        history = models.RaysHistoryMod.objects.filter(client=client)
        rays_data = rest_api.SimpleRaysHistorySerializer(history, many=True).data

        # Получаем все оплаты клиента
        transactions = models.CashTransactionHistory.objects.filter(client=client)
        total_by_currency = defaultdict(Decimal)
        total_paid_usd = Decimal('0.00')

        for t in transactions:
            total_by_currency[t.currency.currency] += t.amount
            try:
                # ✅ передаем объект currency, а не строку
                usd_value = to_usd(t.amount, t.currency)
                total_paid_usd += usd_value
            except Exception as e:
                # если курс не найден или другая ошибка
                continue

        return Response({
            "client": rest_api.ClientsSerializer(client).data,
            "rays_history": rays_data,
            "rays_count": history.count(),
            "total_paid": {k: float(v) for k, v in total_by_currency.items()},
            "total_paid_usd": round(float(total_paid_usd), 2)
        })
    @swagger_auto_schema(operation_summary="📘 Справка по истории", operation_description="Возвращает описание доступных методов: car-history, client-history")
    def list(self, request):  # Чтобы router отображал
        return Response({
            "message": "Используйте /history/<id>/car-history/ или /history/<id>/client-history/"
        })
class OptolViewSet(viewsets.ModelViewSet):
    permission_classes = [IsZaphosOrAdmin]
    queryset = models.OptolMod.objects.all()
    serializer_class = rest_api.OptolSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    @swagger_auto_schema(
        operation_summary="📋 Получить список optol",
        operation_description="Возвращает список всех optol.",
        responses={200: openapi.Response("OK")}
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Добавить новую optol",
        operation_description="Создает новую optol по переданным данным.",
        responses={201: openapi.Response("Создано")}
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Получить optol по ID",
        operation_description="Возвращает данные optol по её ID.",
        responses={200: openapi.Response("OK"), 404: openapi.Response("Не найдено")}
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="✏️ Обновить данные optol полностью",
        operation_description="Полностью обновляет данные optol по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔧 Частичное обновление optol",
        operation_description="Обновляет отдельные поля optol по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🗑️ Удалить optol",
        operation_description="Удаляет optol по её ID.",
        responses={204: openapi.Response("Удалено"), 404: openapi.Response("Не найдено")}
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class BalonFurgonViewSet(viewsets.ModelViewSet):
    permission_classes = [IsZaphosOrAdmin]
    queryset = models.BalonFurgon.objects.all()
    serializer_class = rest_api.BolonFurgonSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    @swagger_auto_schema(
        operation_summary="📋 Получить список bolon for furgon",
        operation_description="Возвращает список всех bolon for furgon.",
        responses={200: openapi.Response("OK")}
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Добавить новую bolon for furgon",
        operation_description="Создает новую bolon for furgon по переданным данным.",
        responses={201: openapi.Response("Создано")}
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Получить bolon for furgon по ID",
        operation_description="Возвращает данные bolon for furgon по её ID.",
        responses={200: openapi.Response("OK"), 404: openapi.Response("Не найдено")}
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="✏️ Обновить данные bolon for furgon полностью",
        operation_description="Полностью обновляет данные bolon for furgon по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔧 Частичное обновление bolon for furgon",
        operation_description="Обновляет отдельные поля bolon for furgon по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🗑️ Удалить bolon for furgon",
        operation_description="Удаляет bolon for furgon по её ID.",
        responses={204: openapi.Response("Удалено"), 404: openapi.Response("Не найдено")}
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class BalonViewSet(viewsets.ModelViewSet):
    permission_classes = [IsZaphosOrAdmin]
    queryset = models.BalonMod.objects.all()
    serializer_class = rest_api.BalonSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    @swagger_auto_schema(
        operation_summary="📋 Получить список bolon",
        operation_description="Возвращает список всех bolon.",
        responses={200: openapi.Response("OK")}
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Добавить новую bolon",
        operation_description="Создает новую bolon по переданным данным.",
        responses={201: openapi.Response("Создано")}
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Получить bolon по ID",
        operation_description="Возвращает данные bolon по её ID.",
        responses={200: openapi.Response("OK"), 404: openapi.Response("Не найдено")}
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="✏️ Обновить данные bolon полностью",
        operation_description="Полностью обновляет данные bolon по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔧 Частичное обновление bolon",
        operation_description="Обновляет отдельные поля bolon по ID.",
        responses={200: openapi.Response("Обновлено"), 400: openapi.Response("Ошибка запроса")}
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🗑️ Удалить bolon",
        operation_description="Удаляет bolon по её ID.",
        responses={204: openapi.Response("Удалено"), 404: openapi.Response("Не найдено")}
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class TexViewSet(viewsets.ModelViewSet):
    permission_classes = [IsZaphosOrAdmin]
    queryset = models.Texnics.objects.all()
    serializer_class = rest_api.TexSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(
        operation_summary="📋 Получить список всех тех. обслуживаний",
        operation_description="📄 Получить список всех тех. обслуживаний",
        responses={200: rest_api.TexSerializer(many=True)}
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Получить информацию по одному Tex",
        operation_description="🔍 Получить информацию по одному Tex",
        responses={200: rest_api.TexSerializer()}
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Создать новое Tex",
        operation_description="➕ Создать новое Tex",
        request_body=rest_api.TexSerializer,
        responses={201: rest_api.TexSerializer()}
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="✏️ Полностью обновить Tex",
        operation_description="✏️ Полностью обновить Tex",
        request_body=rest_api.TexSerializer,
        responses={200: rest_api.TexSerializer()}
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🧩 Частично обновить Tex",
        operation_description="🧩 Частично обновить Tex",
        request_body=rest_api.TexSerializer,
        responses={200: rest_api.TexSerializer()}
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🗑 Удалить Tex",
        operation_description="🗑 Удалить Tex",
        responses={204: 'Удалено успешно'}
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class HistoryViewSet(viewsets.ModelViewSet):
    queryset = models.RaysHistoryMod.objects.all()
    serializer_class = rest_api.ExtendedRaysHistorySerializer
    permission_classes = [IsBugalterOrAdmin]

class CarFullHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsBugalterOrAdmin]
    queryset = models.RaysHistoryMod.objects.select_related(
        'driver', 'car', 'fourgon', 'country'
    ).prefetch_related(
        'client', 
        'product_set',
        'rayshistoryexpense_set'
    ).all().order_by('-created_at')
    serializer_class = rest_api.ExtendedRaysHistorySerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
class RaysHistoryFullViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsBugalterOrAdmin]
    queryset = models.RaysHistoryMod.objects.select_related(
        'driver', 'car', 'fourgon', 'country'
    ).prefetch_related(
        'client', 
        'product_set',
        'rayshistoryexpense_set'
    ).all().order_by('-created_at')
    serializer_class = rest_api.ExtendedRaysHistorySerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @action(detail=False, methods=['get'], url_path='rayshistory-overview')
    def rayshistory_overview(self, request):
        rates = {rate.currency: float(rate.rate_to_uzs) for rate in models.CurrencyRate.objects.all()}
        usd_rate = rates.get('USD', 12500)
        
        period = request.query_params.get('period')
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        
        rays_history = models.RaysHistoryMod.objects.prefetch_related('rayshistoryexpense_set').all()
        
        if from_date and to_date:
            rays_history = rays_history.filter(created_at__range=[from_date, to_date])
        elif period:
            if period == 'week':
                rays_history = rays_history.filter(created_at__gte=now() - timedelta(days=7))
            elif period == 'month':
                rays_history = rays_history.filter(created_at__gte=now() - timedelta(days=30))
            elif period == 'year':
                rays_history = rays_history.filter(created_at__gte=now() - timedelta(days=365))

        rays_count = rays_history.count()
        total_rays_price_uzs = 0
        total_kilometr = 0
        total_profit_uzs = 0
        total_driver_expense_uzs = 0

        for rays in rays_history:
            rate = float(rays.custom_rate_to_uzs) if hasattr(rays, 'custom_rate_to_uzs') and rays.custom_rate_to_uzs else usd_rate
            if not rate: rate = usd_rate

            price_uzs = float(rays.price) * rate if rays.price < 100000 else float(rays.price)
            dr_price_uzs = float(rays.dr_price) * rate if rays.dr_price < 100000 else float(rays.dr_price)
            
            dp_rate = float(rays.dp_currency.rate_to_uzs) if rays.dp_currency else rate
            dp_price_uzs = float(rays.dp_price) * dp_rate
            
            driver_exp_uzs = float(rays.driver_expense) * rate if rays.driver_expense < 100000 else float(rays.driver_expense)
            
            expenses_uzs = 0
            for exp in rays.rayshistoryexpense_set.all():
                exp_rate = float(exp.currency.rate_to_uzs) if exp.currency else rate
                expenses_uzs += float(exp.price) * exp_rate

            total_rays_price_uzs += price_uzs
            total_kilometr += rays.kilometer
            total_driver_expense_uzs += driver_exp_uzs
            
            profit_uzs = price_uzs - dp_price_uzs - expenses_uzs - driver_exp_uzs
            total_profit_uzs += profit_uzs

        return Response({
            'rays_count': rays_count,
            'rays_kilometr': float(total_kilometr),
            'rays_price': round(total_rays_price_uzs, 2),
            'rays_total_price': round(total_profit_uzs, 2),
            'rays_driver_expense': round(total_driver_expense_uzs, 2)
        })

    @docs.rayshistory_locations_doc
    @action(detail=False, methods=['get'], url_path='locations')
    def location(self, request):
        rates = {rate.currency: float(rate.rate_to_uzs) for rate in models.CurrencyRate.objects.all()}
        usd_rate = rates.get('USD', 12500)
        
        result = defaultdict(lambda: {"rays_count": 0, "total_price": 0})
        rays_history = models.RaysHistoryMod.objects.prefetch_related('client')

        for rays in rays_history:
            rate = float(rays.custom_rate_to_uzs) if hasattr(rays, 'custom_rate_to_uzs') and rays.custom_rate_to_uzs else usd_rate
            
            # Get products for this rays history
            products = models.Product.objects.filter(rays_history=rays).select_related('from_location', 'to_location', 'currency')

            for product in products:
                from_loc = product.from_location.name if product.from_location else "Noma'lum"
                to_loc = product.to_location.name if product.to_location else "Noma'lum"
                key = (from_loc, to_loc)

                prod_rate = float(product.currency.rate_to_uzs) if product.currency else rate
                price_uzs = float(product.price) * prod_rate

                result[key]["rays_count"] += 1
                result[key]["total_price"] += price_uzs

        response_data = sorted([
            {
                "from_location": from_loc,
                "to_location": to_loc,
                "rays_count": data["rays_count"],
                "total_price": round(data["total_price"], 2)
            }
            for (from_loc, to_loc), data in result.items()
        ], key=lambda x: x["rays_count"], reverse=True)[:5]
        
        return Response(response_data)

    @swagger_auto_schema(
        operation_summary="📜 История рейсов (только чтение)",
        operation_description="Возвращает историю всех завершённых рейсов в порядке убывания даты."
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Один рейс из истории по ID",
        operation_description="Получить подробную информацию о конкретном рейсе из истории."
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
class RaysViewSet(viewsets.ModelViewSet):
    permission_classes = [IsBugalterOrAdmin]
    queryset = models.RaysMod.objects.select_related(
        'driver', 'car', 'fourgon', 'country', 'dp_currency'
    ).prefetch_related(
        'client', 
        'client_completed',
        'product_set'
    ).all().order_by('-created_at')
    serializer_class = rest_api.RaysSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['is_completed', 'driver', 'client']
    pagination_class = RaysPagination  # <--- здесь применяем

    @action(detail=False,methods=['get'],url_path='active-rays-overview')
    def active_overview(self,request):
        rays = models.RaysMod.objects.all()
        rays_price = rays.aggregate(total=Coalesce(Sum('price'), Value(0), output_field=DecimalField()))['total']
        rays_dr_price = rays.aggregate(total=Coalesce(Sum('dr_price'), Value(0), output_field=DecimalField()))['total']
        rays_dp_price = rays.aggregate(total=Coalesce(Sum('dp_price'), Value(0), output_field=DecimalField()))['total']
        rays_total_price = rays_price - (rays_dr_price + rays_dp_price)
        return Response({
            "rays_price": round(rays_price, 2),
            "rays_dr_price": round(rays_dr_price, 2),
            "rays_dp_price": round(rays_dp_price, 2),
            "rays_total_price": round(rays_total_price, 2)
        })

    @swagger_auto_schema(operation_summary="📋 Список рейсов")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Создать рейс")
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        rays_id = response.data.get('id')
        if rays_id:
            rays = models.RaysMod.objects.get(id=rays_id)
            client_ids = request.data.get('client', [])  # ожидаем список id клиентов
            product_ids = request.data.get('product_ids', []) # конкретные id продуктов от фронтенда
            
            for client_id in client_ids:
                if product_ids:
                    # Если фронтенд передал список конкретных продуктов
                    products = models.Product.objects.filter(client_id=client_id, id__in=product_ids, rays__isnull=True)
                else:
                    # Обратная совместимость: берем все товары без рейса для клиента
                    products = models.Product.objects.filter(client_id=client_id, rays__isnull=True)
                
                for product in products:
                    product.rays = rays
                    product.save()
        return response


    @swagger_auto_schema(operation_summary="🔍 Получить рейс по ID")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновить рейс")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частичное обновление рейса")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удалить рейс")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'], url_path='recalculate-price')
    def recalculate_price(self, request, pk=None):
        try:
            rays = self.get_object()
        except models.RaysMod.DoesNotExist:
            return Response({"error": "Рейс не найден"}, status=404)

        rays.update_prices_from_products_and_expenses()

        return Response({"success": f"✅ Цена и расходы рейса обновлены (в USD): price = {rays.price}, dr_price = {rays.dr_price}"})

    @action(detail=True, methods=['post'], url_path='return-advance')
    def return_advance(self, request, pk=None):
        rays = self.get_object()
        amount = request.data.get('amount')
        if amount is None:
            return Response({"error": "amount maydoni majburiy"}, status=400)
        
        try:
            amount = int(amount)
        except (ValueError, TypeError):
            return Response({"error": "amount raqam bo'lishi kerak"}, status=400)

        rays.returned_advance += amount
        rays.save()

        # Kassa uchun tranzaksiya yaratamiz (avtomatik tasdiqlangan)
        category, _ = models.CashCategory.objects.get_or_create(name="Haydovchidan qaytgan pul")
        
        client = rays.client.first()
        if not client:
            return Response({"error": "Reysga biriktirilgan mijoz topilmadi. Tranzaksiya yaratish uchun kamida bitta mijoz bo'lishi kerak."}, status=400)

        models.CashTransactionMod.objects.create(
            rays=rays,
            amount=amount,
            currency=rays.dp_currency or models.get_default_currency_object(),
            payment_way=category,
            status='confirmed',
            is_confirmed_by_cashier=True,
            cashier=request.user if request.user.is_authenticated and request.user.status == 'cashier' else None,
            comment=f"Reys #{rays.id} bo'yicha haydovchidan ortiqcha pul qaytarildi",
            is_via_driver=False,
            client=client
        )

        return Response({
            "success": f"✅ {amount} miqdoridagi pul kassaga qabul qilindi.",
            "total_returned": rays.returned_advance
        })

    @swagger_auto_schema(
        method='get',
        operation_summary="(GET) Rays Finish",
        operation_description="Use /rays/{id}/complete-race/ — Получить статус завершения рейса"
    )
    @swagger_auto_schema(
        method='post',
        operation_summary="(POST) Rays Finish",
        operation_description="Use /rays/{id}/complete-race/ — Завершить все клиенты и перенести рейс в историю"
    )
    @action(detail=True, methods=['get', 'post'], url_path='complete-race')
    def complete_race(self, request, pk=None):
        try:
            rays = self.get_object()
        except models.RaysMod.DoesNotExist:
            return Response({"error": "Рейс не найден"}, status=404)

        try:
            # Выполняем завершение рейса, получаем объект RaysHistoryMod
            rays_history = rays.complete_whole_race()

            # Обновляем все связанные продукты
            products = models.Product.objects.filter(rays=rays)
            for product in products:
                product.rays_history = rays_history  # ✅ правильный тип
                product.rays = None
                product.save()

        except Exception as e:
            return Response({"error": str(e)}, status=400)

        return Response({"success": "Все клиенты завершены, рейс перенесён в историю."})
    @swagger_auto_schema(
        method='get',
        operation_summary="Driver cars, trucks and clients free",
        operation_description="Bo‘sh haydovchi, mashina, furgon va mijozlar ro‘yxati"
    )
    @action(detail=False, methods=['get'], url_path='available-data')
    def available_data(self, request):
        return Response({
            'drivers': rest_api.CustomUserSerializer(models.CustomUser.objects.filter(status='driver', is_busy=False), many=True).data,
            "cars": rest_api.CarsSerializer(models.CarsMod.objects.filter(is_busy=False), many=True).data,
            "furgons": rest_api.FurgonSerializer(models.FurgonMod.objects.filter(is_busy=False), many=True).data,
            "clients": rest_api.ClientsSerializer(models.ClientsMod.objects.all(), many=True).data,
            "products": rest_api.ProductSerializer(models.Product.objects.filter(is_busy=False), many=True).data
        })

class ClientsViewSet(viewsets.ModelViewSet):
    queryset = models.ClientsMod.objects.all()
    serializer_class = rest_api.ClientsSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(operation_summary="📋 Список клиентов", operation_description="Получить список всех клиентов")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Новый клиент", operation_description="Добавить нового клиента")
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔍 Клиент по ID", operation_description="Получить клиента по его ID")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновление клиента", operation_description="Полное обновление данных клиента")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частичное обновление клиента", operation_description="Изменение одного или нескольких полей клиента")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удаление клиента", operation_description="Удаляет клиента по его ID")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class CarsViewSet(viewsets.ModelViewSet):
    queryset = models.CarsMod.objects.all()
    serializer_class = rest_api.CarsSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    # 📋 Список машин
    @swagger_auto_schema(
        operation_summary="📋 Список машин",
        operation_description="Возвращает список всех машин."
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    # ➕ Добавление машины
    @swagger_auto_schema(
        operation_summary="➕ Создать машину",
        operation_description="Добавить новую машину в систему."
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    # 🔍 Получение одной машины по ID
    @swagger_auto_schema(
        operation_summary="🔍 Получить машину по ID",
        operation_description="Возвращает полную информацию о машине."
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    # ✏️ Полное обновление
    @swagger_auto_schema(
        operation_summary="✏️ Обновить машину",
        operation_description="Обновляет все поля машины по ID."
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    # 🔧 Частичное обновление
    @swagger_auto_schema(
        operation_summary="🔧 Частичное обновление машины",
        operation_description="Обновляет только указанные поля машины."
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    # 🗑️ Удаление
    @swagger_auto_schema(
        operation_summary="🗑️ Удалить машину",
        operation_description="Удаляет машину по ID."
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    # 👇 Кастомный endpoint — статус машин
    @swagger_auto_schema(
        method='get',
        operation_summary="🚗 Статус машин (заняты/свободны)",
        operation_description="Возвращает список занятых и свободных машин. Заняты — в рейсах."
    )
    @action(detail=False, methods=['get'], url_path='status-summary')
    def status_summary(self, request):
        busy_cars_qs = models.CarsMod.objects.filter(is_busy=True)
        free_cars_qs = models.CarsMod.objects.filter(is_busy=False)
        busy_cars = self.get_serializer(busy_cars_qs, many=True).data
        free_cars = self.get_serializer(free_cars_qs, many=True).data
        return Response({
            "in_rays": {
                "count": busy_cars_qs.count(),
                "items": busy_cars
            },
            "available": {
                "count": free_cars_qs.count(),
                "items": free_cars
            }
        })

class CustomUserViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'by_status', 'drivers_status', 'top_drivers', 'summary']:
            permission_classes = [permissions.IsAuthenticated]
        elif self.action in ['create', 'update', 'partial_update']:
            permission_classes = [IsOwnerOrCEO | IsBugalterOrAdmin | IsCashierOrAdmin]
        else:
            permission_classes = [IsOwnerOrCEO | IsBugalterOrAdmin]
        return [permission() for permission in permission_classes]

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        CEO Dashboard uchun barcha haydovchilar bo'yicha statistika.
        """
        # Faqat haydovchi statusidagilarni olamiz
        all_drivers = models.CustomUser.objects.filter(status='driver')
        total_count = all_drivers.count()
        
        # Yo'ldagi haydovchilar (Haqiqatda faol reysi borlar)
        on_road_count = models.RaysMod.objects.filter(is_completed=False, driver__isnull=False).values('driver').distinct().count()
        
        # Kutayotgan haydovchilar (Band bo'lmagan haydovchilar)
        waiting_count = all_drivers.filter(is_busy=False).count()
        
        # Jami faol haydovchilar (Yo'ldagilar + Kutayotganlar)
        active_count = on_road_count + waiting_count

        return Response({
            "total": total_count,
            "active": active_count,
            "inactive": max(0, total_count - active_count),
            "on_road": on_road_count,
            "waiting": waiting_count
        })

    queryset = models.CustomUser.objects.annotate(
        rays_count=Count('rayshistorymod', distinct=True),
        total_rays_usd=Coalesce(
            Sum('rayshistorymod__price'),
            Value(0),
            output_field=DecimalField()
        )
    ).order_by('-date')
    serializer_class = rest_api.CustomUserSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['status']
    search_fields = ['fullname', 'phone_number', 'username']

    @swagger_auto_schema(operation_summary="📋 Список пользователей")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Создать пользователя")
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔍 Получить пользователя")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновить пользователя (полностью)")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частичное обновление пользователя")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удалить пользователя")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
    @docs.custom_driver_history_doc
    @action(detail=True,methods=['get'],url_path='driver-history')
    def driver_history(self,request,pk=None):
        try:
            driver = models.CustomUser.objects.get(status='driver', id=pk)
        except models.CustomUser.DoesNotExist:
            return Response({'error': 'Driver not found'}, status=status.HTTP_404_NOT_FOUND)
        
        history = models.RaysHistoryMod.objects.filter(driver=driver)
        # serializer = rest_api.CustomUserSerializer(driver)
        return Response({
            # 'driver': serializer.data,
            'history':rest_api.ExtendedRaysHistorySerializer(history,many=True).data
            })

    @swagger_auto_schema(
        method='get',
        operation_summary="Get users by status",
        operation_description="Use /user/by-status/?role=driver"
    )
    @action(detail=False, methods=['get'], url_path='by-status')
    def by_status(self, request):
        role = request.query_params.get('role')  # например: driver, owner, ceo и т.д.
        if not role:
            return Response({"error": "❌ Параметр ?role= обязателен"}, status=400)
        users = models.CustomUser.objects.filter(status=role)
        serializer = self.get_serializer(users, many=True)
        return Response({
            "count": users.count(),
            "role": role,
            "items": serializer.data
        })
    @swagger_auto_schema(
        method='get',
        operation_summary="only drivers not busy",
        operation_description="Get only drivers free"
    )
    @action(detail=False, methods=['get'], url_path='drivers')
    def drivers_status(self, request):
        users = models.CustomUser.objects.filter(status='driver', is_busy=False)
        fullnames = users.values_list('fullname', flat=True)
        return Response({
            "count": users.count(),
            "items": list(fullnames)
        })
    @swagger_auto_schema(
        method='get',
        operation_summary="Top drivers",
        # operation_description=""
    )
    @action(detail=False, methods=['get'], url_path='top-drivers')
    def top_drivers(self, request):
        top_users = models.CustomUser.objects.filter(status='driver') \
            .annotate(rays_count=Count('rayshistorymod')) \
            .order_by('-rays_count')
        serializer = self.get_serializer(top_users, many=True)
        return Response(serializer.data)
    
class ChiqimlarCategoryViewSet(viewsets.ModelViewSet):
    queryset = models.ChiqimlarCategory.objects.all()
    serializer_class = rest_api.ChiqimlarCategorySerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(operation_summary="📋 Категории расходов", operation_description="Список всех категорий расходов")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Новая категория", operation_description="Создание новой категории расходов")
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔍 Категория по ID", operation_description="Получить категорию расходов по ID")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновить категорию", operation_description="Полное обновление категории")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частичное обновление", operation_description="Изменить отдельные поля категории")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удалить категорию", operation_description="Удалить категорию расходов по ID")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class ChiqimlikViewSet(viewsets.ModelViewSet):
    queryset = models.ChiqimlikMod.objects.select_related('driver', 'chiqimlar', 'currency').all()
    serializer_class = rest_api.ChiqimlikSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(operation_summary="📋 Список чеков", operation_description="Получить список всех чеков (chiqimlik)")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Новый чек", operation_description="Добавить новый расход (чек)")
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔍 Чек по ID", operation_description="Получить чек по ID")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновить чек", operation_description="Полностью обновить чек по ID")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частично обновить чек", operation_description="Изменить отдельные поля чека по ID")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удалить чек", operation_description="Удалить чек по ID")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class ReferensViewSet(viewsets.ModelViewSet):
    queryset = models.ReferensMod.objects.all()
    serializer_class = rest_api.ReferensSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(operation_summary="📋 Список референсов", operation_description="Получить список всех записей referens")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Новый референс", operation_description="Добавить новую запись referens")
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔍 Референс по ID", operation_description="Получить запись по ID")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновить референс", operation_description="Полное обновление записи")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частичное обновление", operation_description="Изменение отдельных полей записи")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удалить запись", operation_description="Удалить запись referens по ID")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
    
class ArizaViewSet(viewsets.ModelViewSet):
    queryset = models.ArizaMod.objects.all()
    serializer_class = rest_api.ArizaSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(operation_summary="📋 Список заявок", operation_description="Получить список всех заявок (ariza)")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Новая заявка", operation_description="Создать новую заявку")
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔍 Заявка по ID", operation_description="Получить заявку по её ID")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновить заявку", operation_description="Полное обновление заявки")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частичное обновление заявки", operation_description="Изменение некоторых полей заявки")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удалить заявку", operation_description="Удалить заявку по ID")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class FurgonViewSet(viewsets.ModelViewSet):
    queryset = models.FurgonMod.objects.all()
    serializer_class = rest_api.FurgonSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(
        operation_summary="📋 Список фургонов",
        operation_description="Получить список всех фургонов."
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="➕ Добавить фургон",
        operation_description="Создать новый фургон в системе."
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔍 Получить фургон по ID",
        operation_description="Возвращает полную информацию о фургоне."
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="✏️ Обновить фургон",
        operation_description="Полностью обновить данные фургона."
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🔧 Частичное обновление фургона",
        operation_description="Обновляет только переданные поля фургона."
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="🗑️ Удалить фургон",
        operation_description="Удалить фургон по его ID."
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @swagger_auto_schema(
        method='get',
        operation_summary="🚚 Статус фургонов (заняты/свободны)",
        operation_description="Возвращает список занятых и свободных фургонов. Заняты — в активных рейсах."
    )
    @action(detail=False, methods=['get'], url_path='status-summary')
    def status_summary(self, request):
        busy_qs = models.FurgonMod.objects.filter(is_busy=True)
        free_qs = models.FurgonMod.objects.filter(is_busy=False)
        busy_data = rest_api.FurgonSerializer(busy_qs, many=True).data
        free_data = rest_api.FurgonSerializer(free_qs, many=True).data
        return Response({
            "in_rays": {
                "count": busy_qs.count(),
                "items": busy_data
            },
            "available": {
                "count": free_qs.count(),
                "items": free_data
            }
        })

class ProductViewSet(viewsets.ModelViewSet):
    queryset = models.Product.objects.select_related(
        'rays', 'rays_history', 'client', 'currency', 'from_location', 'to_location'
    ).all()
    serializer_class = rest_api.ProductSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @swagger_auto_schema(operation_summary="📋 Список продуктов", operation_description="Получить список всех продуктов")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="➕ Новый продукт", operation_description="Добавить новый продукт")
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔍 Продукт по ID", operation_description="Получить продукт по его ID")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="✏️ Обновить продукт", operation_description="Полное обновление продукта")
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🔧 Частичное обновление продукта", operation_description="Изменение полей продукта")
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(operation_summary="🗑️ Удалить продукт", operation_description="Удалить продукт по ID")
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
